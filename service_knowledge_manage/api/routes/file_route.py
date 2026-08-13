#!/usr/bin/env python
"""
@File         :file_route.py
@Description  :
@Author       :QiangQu
@Date         :2024/09/03 17:25:51
"""

import asyncio
import json
import os
import re
import subprocess
from datetime import datetime
from typing import Optional

import aiofiles
import openpyxl
import pandas as pd
from bson import ObjectId
from fastapi import APIRouter, BackgroundTasks, Body, Depends, File, Request, UploadFile
from fastapi.responses import FileResponse, Response
from pymilvus.exceptions import MilvusException
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from sqlalchemy.orm import Session
from starlette.concurrency import run_in_threadpool

from base_configs.mongodb_config import CollectionConfig
from base_utils.mongodb_util import MongodbUtil
from base_utils.mysql_util import SessionLocal
from base_utils.ret_util import RetUtil
from service_agent_manage.service.agent_service import AgentService
from service_knowledge_manage.api.routes.knowledge_hub_route import knwolege_retrieval
from service_knowledge_manage.entity.file_entity import ChunkEditInfo, ChunkQueryInfo, FileQueryInfo
from service_knowledge_manage.entity.knowledge_hub_entity import KnowledgeRetrivalInfo
from service_knowledge_manage.service.file_processing_service import file_processing_service
from service_knowledge_manage.service.knowledge_file_service import (
    Knowledge_File_service,
)
from service_knowledge_manage.service.parse_service import FileParseService
from service_knowledge_manage.service.util.file_progress import get_progress
from service_usr_manage.service.snow_util import generate_unique_id

router = APIRouter()

from pathlib import Path

from loguru import logger

from base_utils.minio_util import MinIoUtil


# logger = loguru logger (auto-migrated)
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@router.post("/get_multimodal_model", summary="获取运行中的多模态模型")
async def get_multimodal_model() -> Response:
    try:
        internal_model, external_model = await Knowledge_File_service.get_multimodal_model()

        result = []
        if internal_model["children"]:
            result.append(internal_model)
        if external_model["children"]:
            result.append(external_model)
        return RetUtil.response_ok(result)

    except Exception as e:
        logger.exception("获取运行中的多模态模型失败", str(e))
        return RetUtil.response_error(message="获取运行中的多模态模型失败")


@router.post("/metadata_info_edit", summary="元数据信息编辑")
async def metadata_info_edit(params: ChunkEditInfo):
    try:
        result = await Knowledge_File_service.metadata_info_edit(params.id, params.new_content, params.file_id)
        return RetUtil.response_ok(result)
    except Exception as e:
        logger.exception("元数据信息编辑失败", str(e))
        return RetUtil.response_error(message="元数据信息编辑失败")


@router.post("/delete_metadata_info", summary="删除元数据表格、图片描述信息")
async def edit_metadata_info(
    file_id: str = Body(..., embed=True, description="文件id"), id: str = Body(..., embed=True, description="节点id")
) -> Response:
    try:
        result = await Knowledge_File_service.delete_metadata_info(file_id, id)
        return RetUtil.response_ok(result)
    except Exception as e:
        logger.exception("删除元数据表格、图片描述信息失败", str(e))
        return RetUtil.response_error(message="删除元数据表格、图片描述信息失败")


@router.post("/update_parse_result", summary="修改解析结果")
async def update_parse_result(
    file_id: str = Body(..., embed=True, description="文件id"),
    data: list = Body(..., embed=True, description="修改id、内容"),
):
    """
    文件上传解析后，保存修改解析结果
    :param file_id: 文件id
    :param new_content: 修改内容
    :param content_id: 修改内容id
    """
    try:
        result = await Knowledge_File_service.update_parse_result(file_id, data)
        return RetUtil.response_ok(result)
    except Exception as e:
        logger.exception("修改解析结果失败", str(e))
        return RetUtil.response_error(message="修改解析结果失败")


@router.post("/query_parse_result_by_id", summary="通过文件ID查询解析结果")
async def update_parse_result(file_id: str = Body(..., embed=True, description="文件id")):
    try:
        parse_result = MongodbUtil.query_doc_by_id(CollectionConfig.FILE_PARSE_RESULT, file_id).get(
            "parse_result", None
        )
        return RetUtil.response_ok(parse_result)
    except Exception as e:
        logger.exception("查询解析结果失败", str(e))


@router.post("/file_upload_parsing", summary="上传文件并解析")
async def file_upload_parsing(request: Request, file_obj: UploadFile = File(..., description="上传文件")) -> Response:
    try:
        local_path = ""
        file_parse = FileParseService()
        file_name = file_obj.filename
        logger.info(f"文件名称:{file_name}")
        local_path = f"{Path(__file__).parents[3]}/upload/{file_name}"
        # 异步写入文件内容
        async with aiofiles.open(local_path, "wb") as temp_file:
            content = await file_obj.read()
            await temp_file.write(content)

        class SimpleRequest:
            def __init__(self, data):
                self.__dict__.update(data or {})

        parse_content = ""
        mock_request = SimpleRequest(request)
        multimodal_id, model_uid, api_url, api_key, is_external = await Knowledge_File_service.get_multimodal_info()
        # 2.文档解析
        docs = await file_parse.parse_file(
            local_path,
            "",
            mock_request,
            file_name=file_name,
            file_id="",
            preview_mode=False,
            multimodal_id=multimodal_id,
            is_header_config=False,
            start_line=0,
            end_line=0,
            header_merge_method="all",
            is_content_merge=False,
            is_save_image=False,
            is_preview=False,
            model_uid=model_uid,
            api_url=api_url,
            api_key=api_key,
            is_external=is_external,
        )
        content_list = (
            docs["results"]["content_list"]
            if not file_name.endswith((".xlsx", ".xls", ".csv"))
            else docs["results"]["content_list"]["result"]
        )
        for node in content_list:
            if node.get("text", None):
                parse_content += node["text"]
                parse_content += "\n"

        return RetUtil.response_ok(parse_content)
    except Exception as e:
        logger.exception("上传文件并解析失败", str(e))
        return RetUtil.response_error(message="上传文件并解析失败")
    finally:
        if local_path:
            if os.path.exists(local_path):
                os.remove(local_path)


@router.post("/file_upload_mutltimode", summary="上传文件")
async def file_upload_mutltimode(
    chat_request: Request,
    file_obj: UploadFile = File(..., description="文件"),
) -> Response:
    try:
        account_id = chat_request.state.account_id
        file_name = file_obj.filename
        logger.info(f"文件名称:{file_name}")
        remote_path = await Knowledge_File_service.get_upload_file_multimode(file_name, file_obj, account_id)

        return RetUtil.response_ok({"remote_path": remote_path})
    except Exception as e:
        logger.exception("上传文件失败", str(e))
        return RetUtil.response_error(message="上传文件失败")


@router.post("/call_knowledge_create_document", summary="文件上传")
async def call_knowledge_create_document(
    request: Request,
    db: Session = Depends(get_db),
    id: str = Body(..., description="知识库id", embed=True),
    chunk_method: str = Body(..., examples=["RecursiveCharacterTextSplitter"], description="切片方式"),
    file_obj: UploadFile = File(..., description="上传文件"),
    chunk_size: int = Body(500, description="文本块大小", examples=[500]),
    chunk_overlap: int = Body(50, description="文本块重叠大小", examples=[50]),
    separator: list[str] = Body(["\n"], description="文本分隔符"),
    is_generate: bool = Body(False, description="是否生成问答对", examples=[False]),
    preview_mode: bool = Body(False, description="是否仅预览"),
    multimodal_id: str = Body("", description="多模态模型id", embed=True),
    is_header_config: bool = Body(False, description="是否配置表头", embed=True),
    start_line: int = Body(0, description="开始行", embed=True),
    end_line: int = Body(0, description="结束行", embed=True),
    header_merge_method: str = Body("", description="表头拼接方式", embed=True),
    is_content_merge: bool = Body(False, description="是否拼接内容", embed=True),
    is_save_image: bool = Body(False, description="是否保存图片", embed=True),
    use_force_separator: bool = Body(False, description="是否使用强制分隔符", embed=False),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    chunk_type: str = Body("", description="切分类型", examples=["tradition"]),
    sub_chunk_size: Optional[int] = Body(500, description="子块文本块大小", examples=["500"]),
    sub_separator: Optional[list] = Body([], description="子块文本分隔符"),
) -> Response:
    try:
        file_id = ""
        if file_obj.filename.split(".")[-1] not in [
            "png",
            "jpg",
            "xlsx",
            "xls",
            "xlsd",
            "docx",
            "pdf",
            "txt",
            "ppt",
            "md",
            "csv",
            "pptx",
            "html",
            "doc",
        ]:
            return RetUtil.response_error(message="文件格式不符合标准，请重新上传")
        if chunk_overlap >= chunk_size:
            return RetUtil.response_error(message="分段重叠长度不能大于分段最大长度")
        if separator == []:
            separator = ["\n"]
        unicode_escape_pattern = re.compile(r"\\u[0-9a-fA-F]{4}|\\x[0-9a-fA-F]{2}|\\[ntrbfv]")
        for i in range(len(separator)):
            if unicode_escape_pattern.search(separator[i]):
                separator[i] = separator[i].encode("utf-8").decode("unicode_escape")
            else:
                pass
        logger.info(f"去除转义字符后的文本分隔符:{separator}")
        account_id = request.state.account_id
        is_own_knowledge = await Knowledge_File_service.is_own_knowledge(knowledge_id=id, account_id=account_id, db=db)
        if not is_own_knowledge:
            return RetUtil.response_error(message="数据越权")
        result = MongodbUtil.query_docs_by_condition(
            CollectionConfig.KB_COLLECTION, search_condition={"_id": ObjectId(id)}
        )
        if len(list(result)) <= 0:
            return RetUtil.response_error(message="用户所属知识库不存在")
        # 上传文件到临时存储
        remote_path = await Knowledge_File_service.get_upload_file_v1(file_obj, file_obj.filename)
        logger.info(f"文件保存的远程路径为{remote_path}")
        if id == "":
            return RetUtil.response_error(message="知识库ID不能为空")
        if chunk_method == "":
            return RetUtil.response_error(message="切块方式不能为空")
        if file_obj == "":
            return RetUtil.response_error(message="上传文件不能为空")
        internal_model, external_model = await Knowledge_File_service.get_multimodal_model()
        if multimodal_id == "":
            for item in internal_model["children"]:
                multimodal_id = item["id"]
                break
            if multimodal_id == "":
                for item in external_model["children"]:
                    multimodal_id = item["id"]
                    break
        if chunk_type == "":
            return RetUtil.response_error(message="切分模式不能为空")

        if chunk_method  in ["CharacterTextSplitter","RecursiveCharacterTextSplitter","SpacyTextSplitter"]:
            if chunk_type!="tradition":
                return RetUtil.response_error(message="切片方式有误!")
        elif chunk_method  in ["parent_by_title", "parent_by_paragraph", "parent_by_page"]:
            if chunk_type!="parent":
                return RetUtil.response_error(message="切片方式有误!")
        else:
            return RetUtil.response_error(message="切片方式有误!")

        if preview_mode == False:
            result = await file_processing_service.save_kb_arrange_info(
                id,
                chunk_method,
                chunk_overlap,
                chunk_size,
                is_generate,
                separator,
                use_force_separator,
                chunk_type,
                sub_chunk_size,
                sub_separator,
            )
            if result == False:
                return RetUtil.response_error(message="切片类型与数据库不符!")


        # 检查文件是否已存在
        if len(
            list(
                MongodbUtil.query_docs_by_condition(
                    collection_name=CollectionConfig.UPLOAD_FILE_INFO_COLLECTION,
                    search_condition={
                        "knowledge_id": id,
                        "file_name": file_obj.filename,
                    },
                )
            )
        ):
            return RetUtil.response_error(message="知识库中已经存在该文件")
        # 使用新的文件处理服务
        await file_processing_service.process_file_upload_batch(
            knowledge_id=id,
            remote_paths=[remote_path],
            delete_files=[],
            chunk_method=chunk_method,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            separator=separator,
            is_generate=is_generate,
            request=request,
            preview_mode=preview_mode,
            multimodal_id=multimodal_id,
            is_header_config=is_header_config,
            start_line=start_line,
            end_line=end_line,
            header_merge_method=header_merge_method,
            is_content_merge=is_content_merge,
            is_save_image=is_save_image,
            background_tasks=background_tasks,
            use_force_separator=use_force_separator,
            chunk_type=chunk_type,
            sub_chunk_size=sub_chunk_size,
            sub_separator=sub_separator,
        )
        file_info = MongodbUtil.query_docs_by_condition(
            collection_name=CollectionConfig.UPLOAD_FILE_INFO_COLLECTION,
            search_condition={"knowledge_id": id, "file_name": file_obj.filename},
        )
        for item in file_info:
            file_id = item.get("_id", "")
        return RetUtil.response_ok(data={"document_id": file_id})
    except Exception as e:
        logger.exception("文件上传失败", str(e))
        return RetUtil.response_error(message="文件上传失败")


@router.post("/call_knowledge_retrieval", summary="知识库检索")
async def call_knowledge_retrieval(
    request: Request,
    db: Session = Depends(get_db),
    params: KnowledgeRetrivalInfo = Body(..., description="知识库检索参数"),
) -> Response:
    try:
        # 参数验证
        if not params.id:
            return RetUtil.response_error(message="知识库ID不能为空")

        if params.recall_num <= 0:
            return RetUtil.response_error(message="召回切片数量不得小于0")

        # 验证重排模型
        model_name = ""
        if params.rerank_id:
            if not re.match(r"^[a-fA-F0-9]{24}$", params.rerank_id):
                return RetUtil.response_error(message="重排模型id不符合规范")

            try:
                result = list(
                    MongodbUtil.query_docs_by_condition(
                        collection_name=CollectionConfig.MODEL_RUN_COLLECTION,
                        search_condition={"_id": ObjectId(params.rerank_id)},
                    )
                )
                if not result:
                    return RetUtil.response_error(message="重排模型id不存在")
                model_name = result[0].get("model_uid", "")
            except Exception as e:
                logger.exception(f"查询重排模型失败: {e}")
                return RetUtil.response_error(message="重排模型验证失败")

        # 验证知识库访问权限
        account_id = request.state.account_id
        try:
            is_own_knowledge = await Knowledge_File_service.is_own_knowledge(
                knowledge_id=params.id, account_id=account_id, db=db
            )
            if not is_own_knowledge:
                return RetUtil.response_error(message="数据越权")
        except Exception as e:
            logger.exception(f"验证知识库权限失败: {e}")
            return RetUtil.response_error(message="权限验证失败")

        # 更新重排模型名称
        params.rerank_model = model_name

        # 调用知识检索服务
        result = await knwolege_retrieval(params)
        response_body = result.body
        response_str = response_body.decode()
        result_data = json.loads(response_str)

        if result_data["status"]:
            logger.info(f"知识库检索成功 - 返回{len(result_data['data']['results'])}条结果")
            return RetUtil.response_ok(result_data["data"]["results"])
        else:
            logger.warning(f"知识库检索失败: {result_data.get('message', '未知错误')}")
            return RetUtil.response_error(message=str(result_data["message"]))

    except MilvusException as e:
        logger.exception(f"Milvus向量数据库异常: {e}")
        return RetUtil.response_error(message="向量检索服务异常")
    except Exception as e:
        logger.exception(f"知识库检索系统异常: {e}")
        return RetUtil.response_error(message="知识库检索失败")


@router.post("/call_knowledge_file_status", summary="入库状态")
async def call_knowledge_file_status(
    request: Request,
    db: Session = Depends(get_db),
    id: str = Body(..., description="知识库id", embed=True),
    file_name: str = Body(..., description="文件名称"),
) -> Response:
    try:
        input_info = {"id": id, "file_name": file_name}
        # logger.info(f"外部接口查看文件入库状态入参信息:{input_info}")
        account_id = request.state.account_id
        is_own_knowledge = await Knowledge_File_service.is_own_knowledge(knowledge_id=id, account_id=account_id, db=db)
        if not is_own_knowledge:
            return RetUtil.response_error(message="数据越权")

        result = MongodbUtil.query_docs_by_condition(
            CollectionConfig.UPLOAD_FILE_INFO_COLLECTION,
            search_condition={"knowledge_id": id, "file_name": file_name},
        )
        status = ""
        for i in result:
            status = str(i["status"])
        if status != "":
            logger.info("入库状态查询成功")
            return RetUtil.response_ok({"status": status})
        else:
            return RetUtil.response_error(message="知识库或文件名称不存在")
    except Exception as e:
        logger.exception("入库状态查询失败", str(e))
        return RetUtil.response_error(message="入库状态查询失败")


@router.post("/call_knowledge_file_list", summary="知识库文件列表")
async def call_knowledge_file_list(
    request: Request,
    db: Session = Depends(get_db),
    id: str = Body(..., description="知识库id", embed=True),
    file_name: Optional[str] = Body("", embed=True, examples=["file_name"], description="文件名称"),
    page: int = Body(..., embed=True, examples=[1], description="页码"),
    page_size: int = Body(..., embed=True, examples=[1], description="分页大小"),
) -> Response:
    try:
        input_info = {"id": id, "file_name": file_name, "page": page, "page_size": page_size}
        if id == "":
            return RetUtil.response_error(message="知识库ID不能为空")
        if page < 0:
            return RetUtil.response_error(message="页码数量不能为负数")
        if page == "":
            return RetUtil.response_error(message="页码数量不能为空")
        if page_size < 0:
            return RetUtil.response_error(message="分页大小不能为负数")
        if page_size == "":
            return RetUtil.response_error(message="分页大小不能为空")

        account_id = request.state.account_id
        is_own_knowledge = await Knowledge_File_service.is_own_knowledge(knowledge_id=id, account_id=account_id, db=db)
        if not is_own_knowledge:
            return RetUtil.response_error(message="数据越权")

        result = MongodbUtil.query_docs_by_condition(
            collection_name=CollectionConfig.KB_COLLECTION,
            search_condition={"_id": ObjectId(id)},
        )
        if len(list(result)) == 0:
            return RetUtil.response_error(message="知识库不存在")

        result = await file_query_page(FileQueryInfo(id=id, file_name=file_name, page=page, page_size=page_size))
        response_body = result.body
        response_str = response_body.decode()
        result = json.loads(response_str)
        result_list = result["data"]["result"]

        if result["status"]:
            logger.info("查询知识库文件列表成功")
            return RetUtil.response_ok(result["data"]["result"])
        else:
            return RetUtil.response_error(message=str(result["message"]))

    except Exception as e:
        logger.exception("查询知识库文件列表失败", str(e))
        return RetUtil.response_error(message="查询知识库文件列表失败")


@router.post("/call_delete_knowledge_file", summary="知识库文件删除")
async def call_delete_knowledge_file(
    request: Request,
    db: Session = Depends(get_db),
    id: str = Body(..., description="知识库id", embed=True),
    file_id: list = Body(..., embed=True, examples=["file_name"], description="文件名称"),
) -> Response:
    try:
        if id == "":
            return RetUtil.response_error(message="知识库ID不能为空")
        if file_id == "" or file_id == []:
            return RetUtil.response_error(message="文件ID不能为空")

        account_id = request.state.account_id
        is_own_knowledge = await Knowledge_File_service.is_own_knowledge(knowledge_id=id, account_id=account_id, db=db)
        if not is_own_knowledge:
            return RetUtil.response_error(message="数据越权")

        result = MongodbUtil.query_docs_by_condition(
            collection_name=CollectionConfig.KB_COLLECTION,
            search_condition={"_id": ObjectId(id)},
        )
        if len(list(result)) == 0:
            return RetUtil.response_error(message="知识库不存在")
        for item_id in file_id:
            result = MongodbUtil.query_docs_by_condition(
                collection_name=CollectionConfig.UPLOAD_FILE_INFO_COLLECTION,
                search_condition={"knowledge_id": id, "_id": item_id},
            )
            if len(list(result)) == 0:
                return RetUtil.response_error(message=f"知识库文件  {item_id}  不存在")

        result = await file_delete(id=id, file_id=file_id)
        response_body = result.body
        response_str = response_body.decode()
        result = json.loads(response_str)

        if result["status"]:
            return RetUtil.response_ok("知识库文件删除成功")
        else:
            return RetUtil.response_error(message=str(result["message"]))

    except Exception as e:
        logger.exception("知识库文件删除失败", str(e))
        return RetUtil.response_error(message="知识库文件删除失败")


@router.post("/call_query_knowledge_chunk", summary="查询知识库切片的外部接口")
async def call_query_knowledge_chunk(
    request: Request,
    db: Session = Depends(get_db),
    id: str = Body(..., description="知识库id", embed=True),
    file_name: Optional[list] = Body([], embed=True, examples=["file_name"], description="文件名称列表"),
    page: int = Body(..., embed=True, examples=[1], description="页码"),
    page_size: int = Body(..., embed=True, examples=[1], description="分页大小"),
    filter_condition: Optional[str] = Body(None, embed=True, description="切片内容模糊筛选条件"),
) -> Response:
    try:
        # chunk_result_query的外部接口
        input_info = {"id": id, "file_name": file_name, "page": page, "page_size": page_size, "filter_condition": filter_condition}
        id_list = []

        if id == "":
            return RetUtil.response_error(message="知识库ID不能为空")
        if page < 0:
            return RetUtil.response_error(message="页码数量不能为负数")
        if page == "":
            return RetUtil.response_error(message="页码数量不能为空")
        if page_size < 0:
            return RetUtil.response_error(message="分页大小不能为负数")
        if page_size == "":
            return RetUtil.response_error(message="分页大小不能为空")

        account_id = request.state.account_id
        is_own_knowledge = await Knowledge_File_service.is_own_knowledge(knowledge_id=id, account_id=account_id, db=db)
        if not is_own_knowledge:
            return RetUtil.response_error(message="数据越权")

        result = MongodbUtil.query_docs_by_condition(
            collection_name=CollectionConfig.KB_COLLECTION,
            search_condition={"_id": ObjectId(id)},
        )
        if len(list(result)) == 0:
            return RetUtil.response_error(message="知识库不存在")
        if file_name:
            for file in file_name:
                result = MongodbUtil.query_docs_by_condition(
                    collection_name=CollectionConfig.UPLOAD_FILE_INFO_COLLECTION,
                    search_condition={"knowledge_id": id, "file_name": file},
                )
                for i in result:
                    id_list.append(i["_id"])

        result = await chunk_result_query(ChunkQueryInfo(id=id, file_id=id_list, page=page, page_size=page_size, filter_condition=filter_condition))
        response_body = result.body
        response_str = response_body.decode()
        result = json.loads(response_str)
        if result["status"]:
            logger.info("外部接口查询知识库切片信息成功")
            # return RetUtil.response_ok({"total": len_result, "result": result, "chunk_type": chunk_type})
            return RetUtil.response_ok({"total": result["data"]["total"], "result": result["data"]["result"], "chunk_type": result["data"]["chunk_type"]})
            # return RetUtil.response_ok(result["data"]["result"])
        else:
            return RetUtil.response_error(message=str(result["message"]))

    except Exception as e:
        logger.exception("查询知识库切片失败", str(e))
        return RetUtil.response_error(message="查询知识库切片失败")


@router.post("/file_obj_preview", summary="上传文件预览")
async def file_obj_preview(
    file_obj: UploadFile = File(..., description="文件"),
    background_tasks: BackgroundTasks = BackgroundTasks(),
) -> Response:
    try:
        # 获取保存文件夹、文件名称与文件名（不带扩展）
        upload_path = Path(__file__).parents[3] / "upload"
        file_name = file_obj.filename
        file_name_without_extension = Path(file_name).stem

        # 定义保存路径：本地路径、输出路径、最终路径
        unique_id = generate_unique_id("FILE_", datacenter_id=1, worker_id=1)
        local_path = f"{upload_path}/{unique_id}$$${file_name}"
        final_path = ""
        output_path = f"{upload_path}/{unique_id}$$${file_name_without_extension}.pdf"
        local_path = os.path.abspath(local_path).replace("\\", "/")
        output_path = os.path.abspath(output_path).replace("\\", "/")

        def convert_to_pdf(local_path, upload_path):
            try:
                subprocess.run(
                    [
                        "soffice",
                        "--headless",
                        "--convert-to",
                        "pdf",
                        "--outdir",
                        upload_path,
                        local_path,
                    ],
                    check=True,
                )
            except subprocess.CalledProcessError as e:
                logger.exception("转换pdf文件出错", str(e))
                raise

        # word文件处理 先读取doc/docx/txt文件内容，将其保存为docx文件，在使用docx2pdf的convert方法将文件转换为pdf类型
        if file_obj.filename.endswith(".docx") or file_obj.filename.endswith(".doc"):
            count = 0
            while count < 3:
                try:
                    if file_name.endswith(".doc") or file_name.endswith(".txt"):
                        final_path = f"{upload_path}/{unique_id}$$${file_name[:-4]}.docx"
                        final_path = os.path.abspath(final_path).replace("\\", "/")
                    else:
                        final_path = local_path
                    async with aiofiles.open(final_path, "wb") as temp_file:
                        content = await file_obj.read()
                        await temp_file.write(content)
                        await asyncio.sleep(1)
                    convert_to_pdf(final_path, upload_path)
                    logger.info("转换docx文件成功")
                    break
                except Exception as e:
                    logger.exception(f"转换docx文件失败,失败次数：第{count}次", str(e))
                    count += 1
            if count == 3:
                return RetUtil.response_error(message="转换docx文件失败")

        # excel文件处理 先使用pandas读取xls/csv文件内容，将其转换为xlsx文件，再使用reportlab库将excel文件内容写入pdf
        elif (
            file_obj.filename.endswith(".xlsx")
            or file_obj.filename.endswith(".xls")
            or file_obj.filename.endswith(".csv")
        ):
            try:
                async with aiofiles.open(local_path, "wb") as temp_file:
                    content = await file_obj.read()
                    await temp_file.write(content)
                if file_name.endswith(".csv"):
                    try:
                        df = pd.read_csv(local_path)
                    except:
                        df = pd.read_csv(local_path, encoding="gbk")
                    df = df.fillna(method="ffill")
                else:
                    try:
                        df = pd.read_excel(local_path)
                    except:
                        df = pd.read_excel(local_path)
                    df = df.fillna(method="ffill")
                final_path = local_path + ".xlsx"
                final_path = os.path.abspath(final_path).replace("\\", "/")
                df.to_excel(final_path, index=False, header=True)
            except Exception as e:
                logger.exception("转换xlsx文件失败", str(e))
                return RetUtil.response_error(message="转换excel文件失败")

            # 设置文件格式，将excel文件内容写入pdf
            font_path = Path(__file__).parents[3] / "Font/simhei.ttf"
            pdfmetrics.registerFont(TTFont("simsun", font_path))
            wb = openpyxl.load_workbook(final_path)
            sheet = wb.active
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
            for row in data:
                for col_index, cell in enumerate(row):
                    if isinstance(cell, str) and len(cell) > 100:
                        wrapped_text = "\n".join(
                            [chunk for chunk in [cell[i : i + 100] for i in range(0, len(cell), 100)] if chunk]
                        )
                        row[col_index] = wrapped_text
            max_col_widths = [max([len(str(row[i])) for row in data]) for i in range(len(data[0]))]
            max_col_widths = [min(i, 100) for i in max_col_widths]
            table_width = sum(max_col_widths) * 8
            table_height = len(data) * 30
            doc = SimpleDocTemplate(output_path, pagesize=(table_width, table_height))
            elements = []
            t = Table(data)
            t_style = TableStyle(
                [
                    ("FONT", (0, 0), (-1, -1), "simsun", 2),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("FONTSIZE", (0, 0), (-1, -1), 5),
                    ("LEADING", (0, 0), (-1, -1), 6),
                    ("AUTOPAD", (0, 0), (-1, -1), 1),
                ]
            )
            t.setStyle(t_style)
            elements.append(t)
            doc.build(elements)

        # pdf文件不做处理，直接返回
        elif file_obj.filename.endswith(".pdf"):
            async with aiofiles.open(local_path, "wb") as temp_file:
                content = await file_obj.read()
                await temp_file.write(content)
            output_path = local_path

        else:
            for i in range(3):
                async with aiofiles.open(local_path, "wb") as temp_file:
                    content = await file_obj.read()
                    await temp_file.write(content)
                    await asyncio.sleep(0.5)
                if os.path.exists(local_path):
                    break
            if not os.path.exists(local_path):
                return RetUtil.response_error(message="文件保存失败，请重试")

            if file_obj.filename.endswith(".jpg") or file_obj.filename.endswith(".png"):
                import fitz  # PyMuPDF

                doc = fitz.open()  # 创建一个空的 PDF 文档
                page = doc.new_page()  # 默认添加 A4 大小的页面
                rect = fitz.Rect(100, 100, 400, 400)  # 定义图片的位置和大小
                page.insert_image(rect, filename=local_path)  # 插入图片

                doc.save(output_path)
                doc.close()
            elif file_name.endswith(".txt"):
                from fpdf import FPDF

                font_path = Path(__file__).parents[3] / "Font/simhei.ttf"
                pdf = FPDF()
                pdf.add_page()
                pdf.add_font("simsun", "", font_path, uni=True)  # uni=True 支持 Unicode 字符
                pdf.set_font("simsun", size=8)

                with open(local_path, encoding="utf-8") as file:
                    lines = file.read().splitlines()  # 使用 splitlines() 读取所有行，自动处理换行符

                for line in lines:
                    pdf.multi_cell(0, 5, txt=line, align="L")  # 写入每一行
                pdf.output(output_path)
            else:
                try:
                    convert_to_pdf(local_path, upload_path)
                except:
                    logger.exception("转换pdf文件失败", str(e))
                    return RetUtil.response_error(message="转换pdf文件失败")

        logger.info(f"本地路径:{local_path},保存路径:{final_path},输出路径:{output_path}")

        def delete_file(file_path: str):
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    logger.info(f"文件已删除: {file_path}")
            except Exception as e:
                logger.exception(f"删除文件失败: {file_path}", str(e))
                raise

        # 将删除文件的任务添加到 background_tasks
        background_tasks.add_task(delete_file, local_path)
        background_tasks.add_task(delete_file, output_path)
        background_tasks.add_task(delete_file, final_path)

        logger.info("上传文件预览成功")

        return FileResponse(
            path=output_path,
            media_type="application/pdf",
            filename=f"{file_name_without_extension}.pdf",
        )

    except Exception as e:
        logger.exception("文件预览失败", str(e))
        return RetUtil.response_error(message="文件预览失败")


@router.post("/file_url_preview", summary="远程文件预览")
async def file_url_preview(
    id: str = Body(..., description="知识库id", embed=True),
    url: str = Body(..., embed=True, description="远程文件地址"),
    background_tasks: BackgroundTasks = BackgroundTasks(),
) -> Response:
    try:
        input_info = {"id": id, "url": url}
        # logger.info(f"远程文件预览入参信息:{input_info}")
        # 下载文件到本地
        upload_path = Path(__file__).parents[3] / "upload"
        file_name = url.split("/")[-1]
        file_name_without_extension = Path(file_name).stem
        unique_id = generate_unique_id("FILE_", datacenter_id=1, worker_id=1)
        local_path = f"{upload_path}/{unique_id}$$${file_name}"
        final_path = ""
        output_path = f"{upload_path}/{unique_id}$$${file_name_without_extension + '.pdf'}"
        local_path = os.path.abspath(local_path).replace("\\", "/")
        output_path = os.path.abspath(output_path).replace("\\", "/")

        if id in url:
            for i in range(3):
                MinIoUtil.download_file(bucket_name="tiance-base", remote_path=url, local_path=local_path)
                await asyncio.sleep(0.5)
                if os.path.exists(local_path):
                    break
        else:
            for i in range(3):
                MinIoUtil.download_file(
                    bucket_name="tiance-base-temp-file-bucket",
                    remote_path=url,
                    local_path=local_path,
                )
                await asyncio.sleep(0.5)
                if os.path.exists(local_path):
                    break

        if not os.path.exists(local_path):
            return RetUtil.response_error(message="文件保存失败，请重试")

        with open(local_path, "rb") as file:
            content = file.read()

        def convert_to_pdf(local_path, upload_path):
            try:
                subprocess.run(
                    [
                        "soffice",
                        "--headless",
                        "--convert-to",
                        "pdf",
                        "--outdir",
                        upload_path,
                        local_path,
                    ],
                    check=True,
                )
            except subprocess.CalledProcessError as e:
                logger.exception("转换pdf文件出错", str(e))
                raise

        # word文件处理 先读取doc/docx/txt文件内容，将其保存为docx文件，在使用docx2pdf的convert方法将文件转换为pdf类型
        if file_name.endswith(".docx") or file_name.endswith(".doc"):
            count = 0
            while count < 3:
                try:
                    if file_name.endswith(".doc") or file_name.endswith(".txt"):
                        final_path = f"{upload_path}/{unique_id}$$${file_name[:-4]}.docx"
                        final_path = os.path.abspath(final_path).replace("\\", "/")
                    else:
                        final_path = local_path
                    async with aiofiles.open(final_path, "wb") as temp_file:
                        await temp_file.write(content)
                    await asyncio.sleep(1)
                    convert_to_pdf(final_path, upload_path)
                    logger.info("转换docx文件成功")
                    break
                except Exception as e:
                    logger.exception(f"转换docx文件失败,失败次数：第{count}次", str(e))
                    count += 1
            if count == 3:
                return RetUtil.response_error(message="转换docx文件失败")

        # excel文件处理 先使用pandas读取xls/csv文件内容，将其转换为xlsx文件，再使用reportlab库将excel文件内容写入pdf
        elif file_name.endswith(".xlsx") or file_name.endswith(".xls") or file_name.endswith(".csv"):
            try:
                async with aiofiles.open(local_path, "wb") as temp_file:
                    await temp_file.write(content)
                if file_name.endswith(".csv"):
                    try:
                        df = pd.read_csv(local_path)
                    except:
                        df = pd.read_csv(local_path, encoding="gbk")
                    df = df.fillna(method="ffill")
                else:
                    try:
                        df = pd.read_excel(local_path)
                    except:
                        df = pd.read_excel(local_path)
                    df = df.fillna(method="ffill")
                final_path = local_path + ".xlsx"
                final_path = os.path.abspath(final_path).replace("\\", "/")
                df.to_excel(final_path, index=False, header=True)
            except Exception as e:
                logger.exception("转换excel文件失败", str(e))
                return RetUtil.response_error(message="转换excel文件失败")
            font_path = Path(__file__).parents[3] / "Font/simhei.ttf"
            pdfmetrics.registerFont(TTFont("simsun", font_path))
            wb = openpyxl.load_workbook(final_path)
            sheet = wb.active
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
            for row in data:
                for col_index, cell in enumerate(row):
                    if isinstance(cell, str) and len(cell) > 100:
                        wrapped_text = "\n".join(
                            [chunk for chunk in [cell[i : i + 100] for i in range(0, len(cell), 100)] if chunk]
                        )
                        row[col_index] = wrapped_text
            max_col_widths = [max([len(str(row[i])) for row in data]) for i in range(len(data[0]))]
            max_col_widths = [min(i, 100) for i in max_col_widths]
            table_width = sum(max_col_widths) * 8
            table_height = len(data) * 30
            doc = SimpleDocTemplate(output_path, pagesize=(table_width, table_height))
            elements = []
            t = Table(data)
            t_style = TableStyle(
                [
                    ("FONT", (0, 0), (-1, -1), "simsun", 2),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ("FONTSIZE", (0, 0), (-1, -1), 5),
                    ("LEADING", (0, 0), (-1, -1), 6),
                    ("AUTOPAD", (0, 0), (-1, -1), 1),
                ]
            )
            t.setStyle(t_style)
            elements.append(t)
            doc.build(elements)

        # pdf文件不做处理，直接返回
        elif file_name.endswith(".pdf"):
            async with aiofiles.open(local_path, "wb") as temp_file:
                await temp_file.write(content)
            output_path = local_path

        else:
            if file_name.endswith(".jpg") or file_name.endswith(".png"):
                import fitz  # PyMuPDF

                doc = fitz.open()  # 创建一个空的 PDF 文档
                page = doc.new_page()  # 默认添加 A4 大小的页面
                rect = fitz.Rect(100, 100, 400, 400)  # 定义图片的位置和大小
                page.insert_image(rect, filename=local_path)  # 插入图片

                doc.save(output_path)
                doc.close()
            elif file_name.endswith(".txt"):
                from fpdf import FPDF

                font_path = Path(__file__).parents[3] / "Font/simhei.ttf"
                pdf = FPDF()
                pdf.add_page()
                pdf.add_font("simsun", "", font_path, uni=True)  # uni=True 支持 Unicode 字符
                pdf.set_font("simsun", size=8)

                with open(local_path, encoding="utf-8") as file:
                    lines = file.read().splitlines()  # 使用 splitlines() 读取所有行，自动处理换行符

                for line in lines:
                    pdf.multi_cell(0, 5, txt=line, align="L")  # 写入每一行

                pdf.output(output_path)

            else:
                convert_to_pdf(local_path, upload_path)
        logger.info(f"本地路径: {local_path}, 保存路径: {final_path}, 输出路径: {output_path}")

        def delete_file(file_path: str):
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    logger.info(f"文件已删除: {file_path}")
            except Exception as e:
                logger.exception(f"删除文件失败: {file_path}", str(e))
                raise

        # 将删除文件的任务添加到 background_tasks
        background_tasks.add_task(delete_file, local_path)
        background_tasks.add_task(delete_file, output_path)
        background_tasks.add_task(delete_file, final_path)

        logger.info("文件预览成功")

        return FileResponse(
            path=output_path,
            media_type="application/pdf",
            filename=f"{file_name_without_extension}.pdf",
        )

    except Exception as e:
        logger.exception("上传文件预览失败", str(e))
        return RetUtil.response_error(message="上传文件预览失败")


@router.post("/knowledge_file_download", summary="知识库文件下载")
async def knowledge_file_download(
    request: Request,
    db: Session = Depends(get_db),
    id: str = Body(..., description="知识库id", embed=True),
    file_name: str = Body(..., embed=True, examples=["file_name"], description="文件名称列表"),
) -> Response:
    try:
        input_info = {"id": id, "file_name": file_name}
        # logger.info(f"知识库文件下载入参信息:{input_info}")
        upload_path = Path(__file__).parents[3] / "upload"

        if id == "":
            return RetUtil.response_error(message="知识库ID不能为空")
        if file_name == "":
            return RetUtil.response_error(message="文件名称不能为空")

        account_id = request.state.account_id
        is_own_knowledge = await Knowledge_File_service.is_own_knowledge(knowledge_id=id, account_id=account_id, db=db)
        if not is_own_knowledge:
            return RetUtil.response_error(message="数据越权")

        kb_name_result = MongodbUtil.query_docs_by_condition(
            collection_name=CollectionConfig.KB_COLLECTION,
            search_condition={"_id": ObjectId(id)},
        )
        if len(list(kb_name_result)) == 0:
            return RetUtil.response_error(message="知识库不存在")
        file_name_result = MongodbUtil.query_docs_by_condition(
            collection_name=CollectionConfig.UPLOAD_FILE_INFO_COLLECTION,
            search_condition={"knowledge_id": id, "file_name": file_name},
        )
        if len(list(file_name_result)) == 0:
            return RetUtil.response_error(message="知识库文件不存在")

        try:
            remote_path = MongodbUtil.query_docs_by_condition(
                collection_name=CollectionConfig.UPLOAD_FILE_INFO_COLLECTION,
                search_condition={"knowledge_id": id, "file_name": file_name},
            )[0]["remote_path"]
            local_path = f"{upload_path}/{id}$$${file_name}"
            await run_in_threadpool(MinIoUtil.download_file, "tiance-base", remote_path, local_path)
        except Exception as e:
            logger.exception("远程下载知识库文件失败", str(e))
            return RetUtil.response_error(message="文件下载失败，请检查下载文件是否成功上传")

        logger.info("知识库文件下载成功")
        return FileResponse(path=local_path, filename=file_name, media_type="application/octet-stream")

    except Exception as e:
        logger.exception("下载知识库文件失败", str(e))
        return RetUtil.response_error(message="下载知识库文件失败")


@router.post("/call_knowledge_create_document_by_zip", summary="上传压缩包")
async def call_knowledge_create_document_by_zip(
    chat_request: Request,
    db: Session = Depends(get_db),
    id: str = Body(..., description="知识库id", embed=True),
    file_obj: UploadFile = File(..., description="压缩包文件"),
    chunk_method: str = Body(
        "RecursiveCharacterTextSplitter",
        examples=["RecursiveCharacterTextSplitter"],
        description="切片方式",
    ),
    chunk_size: int = Body(500, description="文本块大小", examples=["500"]),
    chunk_overlap: int = Body(50, description="文本块重叠大小", examples=["50"]),
    separator: list[str] = Body(["\n"], description="文本分隔符"),
    is_generate: bool = Body(False, description="是否生成问答对", examples=[False]),
    delete_files: list[str] = Body([], description="需要删除的文件ID列表"),
    preview_mode: bool = Body(False, description="是否仅预览"),
    multimodal_id: str = Body("", description="多模态模型id", embed=True),
    is_header_config: bool = Body(False, description="是否配置表头", embed=True),
    start_line: int = Body(0, description="开始行", embed=True),
    end_line: int = Body(0, description="结束行", embed=True),
    header_merge_method: str = Body("", description="表头拼接方式", embed=True),
    is_content_merge: bool = Body(False, description="是否拼接内容", embed=True),
    is_save_image: bool = Body(False, description="是否保存图片", embed=True),
    use_force_separator: bool = Body(False, description="是否使用强制分隔符", embed=False),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    chunk_type: str = Body("", description="切分类型", examples=["tradition"]),
    sub_chunk_size: Optional[int] = Body(500, description="子块文本块大小", examples=["500"]),
    sub_separator: Optional[list] = Body([], description="子块文本分隔符"),
) -> Response:
    """
    上传压缩包并解压文件
    """
    try:
        # 参数校验
        if file_obj.filename.split(".")[-1] not in ["zip", "7z"]:
            return RetUtil.response_error(message="文件格式不符合标准，请上传压缩包文件")

        if chunk_overlap >= chunk_size:
            return RetUtil.response_error(message="分段重叠长度不能大于分段最大长度")

        if separator == []:
            separator = ["\n"]

        # 处理转义字符
        unicode_escape_pattern = re.compile(r"\\u[0-9a-fA-F]{4}|\\x[0-9a-fA-F]{2}|\\[ntrbfv]")
        for i in range(len(separator)):
            if unicode_escape_pattern.search(separator[i]):
                separator[i] = separator[i].encode("utf-8").decode("unicode_escape")
        logger.info(f"去除转义字符后的文本分隔符:{separator}")

        # 权限检查
        account_id = chat_request.state.account_id
        is_own_knowledge = await Knowledge_File_service.is_own_knowledge(knowledge_id=id, account_id=account_id, db=db)
        if not is_own_knowledge:
            return RetUtil.response_error(message="数据越权")

        # 检查知识库是否存在
        result = MongodbUtil.query_docs_by_condition(
            CollectionConfig.KB_COLLECTION, search_condition={"_id": ObjectId(id)}
        )
        if len(list(result)) <= 0:
            return RetUtil.response_error(message="用户所属知识库不存在")

        # 上传并解压压缩包
        results = []
        result = await Knowledge_File_service.upload_and_extract_archive(file_obj, id)
        if len(result["files"]) != 1:
            result["files"] = result["files"][1:]

        # 检查解压后的文件是否重名
        file_names = [os.path.basename(file) for file in result["files"]]
        repeat_files = await Knowledge_File_service.repeat_file_detect(id, file_names)
        if repeat_files:
            return RetUtil.response_error(message=f"以下文件已存在：{', '.join(repeat_files)}，请重命名后上传")
        if chunk_type == "":
            return RetUtil.response_error(message="切分模式不能为空")
        if chunk_method in ["CharacterTextSplitter", "RecursiveCharacterTextSplitter", "SpacyTextSplitter"]:
            if chunk_type != "tradition":
                return RetUtil.response_error(message="切片方式有误!")
        elif chunk_method in ["parent_by_title", "parent_by_paragraph", "parent_by_page"]:
            if chunk_type != "parent":
                return RetUtil.response_error(message="切片方式有误!")
        else:
            return RetUtil.response_error(message="切片方式有误!")


        if preview_mode == False:
            kb_result = await file_processing_service.save_kb_arrange_info(
                id,
                chunk_method,
                chunk_overlap,
                chunk_size,
                is_generate,
                separator,
                use_force_separator,
                chunk_type,
                sub_chunk_size,
                sub_separator,
            )
            if kb_result == False:
                return RetUtil.response_error(message="切片类型与数据库不符!")



        # 上传文件到MinIO并记录信息
        remote_paths = []
        file_id_list = []
        for file in result["files"]:
            file_name = os.path.basename(file)
            if file_name in repeat_files:
                continue
            async with aiofiles.open(file, "rb") as f:
                remote_path = await Knowledge_File_service.get_upload_file_v1(f, file_name)
            remote_paths.append(remote_path)
        # 使用新的文件处理服务
        await file_processing_service.process_file_upload_batch(
            knowledge_id=id,
            remote_paths=remote_paths,
            delete_files=delete_files,
            chunk_method=chunk_method,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            separator=separator,
            is_generate=is_generate,
            request=chat_request,
            preview_mode=preview_mode,
            multimodal_id=multimodal_id,
            is_header_config=is_header_config,
            start_line=start_line,
            end_line=end_line,
            header_merge_method=header_merge_method,
            is_content_merge=is_content_merge,
            is_save_image=is_save_image,
            use_force_separator=use_force_separator,
            background_tasks=background_tasks,
            chunk_type=chunk_type,
            sub_chunk_size=sub_chunk_size,
            sub_separator=sub_separator,
        )

        logger.info("压缩包上传成功")
        return RetUtil.response_ok(f"压缩包上传成功，{remote_paths}文件正在解析和向量入库中")

    except Exception as e:
        logger.exception("上传压缩包失败", str(e))
        return RetUtil.response_error(message="上传压缩包失败")


@router.post("/call_upload_parsing", summary="文件解析外部接口")
async def call_upload_parsing(
    request: Request,
    file_obj: UploadFile = File(..., description="上传文件"),
) -> Response:
    try:
        file_extension = os.path.splitext(file_obj.filename)[1].lower()
        ALLOWED_EXTENSIONS = [
            "png",
            "jpg",
            "xlsx",
            "xls",
            "xslx",
            "doc",
            "docx",
            "pdf",
            "txt",
            "ppt",
            "md",
            "csv",
            "pptx",
            "html",
        ]
        if file_extension not in [f".{ext}" for ext in ALLOWED_EXTENSIONS]:
            return RetUtil.response_error(
                data=f"不支持的文件类型。允许的文件类型包括：{', '.join([f'.{ext}' for ext in ALLOWED_EXTENSIONS])}"
            )
        file_size = len(await file_obj.read())
        # 重置文件指针到开头，以便后续处理
        await file_obj.seek(0)
        # 检查文件大小是否超过50MB
        if file_size > 50 * 1024 * 1024:  # 50MB
            return RetUtil.response_error(data="文件大小只允许50MB")
        result = await file_upload_parsing(request, file_obj)
        response_body = result.body
        response_str = response_body.decode()
        result = json.loads(response_str)

        if result["status"]:
            logger.info("外部接口文件解析成功")
            return RetUtil.response_ok(data=result["data"])

        else:
            return RetUtil.response_error(message=str(result["message"]))

    except Exception as e:
        logger.exception("外部接口文件解析失败", str(e))
        return RetUtil.response_error(message="文件解析失败")


@router.post("/file_upload_1_v2", summary="上传文件")
async def file_upload_1_v2(
    id: str = Body(..., description="知识库id", embed=True),
    file_obj: UploadFile = File(..., description="文件"),
) -> Response:
    """
    第一步：上传文件到临时存储
    """
    try:
        file_name = file_obj.filename
        logger.info(f"开始上传文件: {file_name} 到知识库: {id}")

        remote_path = await Knowledge_File_service.get_upload_file_v1(file_obj, file_name)

        logger.info(f"文件上传成功，远程路径: {remote_path}")
        return RetUtil.response_ok({"remote_path": remote_path})

    except Exception as e:
        logger.exception(f"上传知识库文件失败: {file_obj.filename if file_obj else 'unknown'}", str(e))
        return RetUtil.response_error(message="上传知识库文件失败")


@router.post("/file_upload_2_v2", summary="切片解析")
async def file_upload_2_v2(
    request: Request,
    id: str = Body(..., description="知识库id", embed=True),
    remote_paths: list[str] = Body(..., description="文件远程路径列表", embed=True),
    chunk_method: str = Body("RecursiveCharacterTextSplitter", description="切片方式", embed=True),
    chunk_size: int = Body(500, description="文本块大小", embed=True),
    chunk_overlap: int = Body(50, description="文本块重叠大小", embed=True),
    separator: list = Body("", description="文本分隔符", embed=True),
    is_generate: bool = Body(False, description="是否生成问答对", embed=True),
    preview_mode: bool = Body(False, description="文件解析模式(False = 快速模式, True = 解析预览模式)", embed=True),
    multimodal_id: str = Body("", description="多模态模型id", embed=True),
    is_header_config: bool = Body(False, description="是否配置表头", embed=True),
    start_line: int = Body(0, description="开始行", embed=True),
    end_line: int = Body(0, description="结束行", embed=True),
    header_merge_method: str = Body("", description="表头拼接方式", embed=True),
    is_content_merge: bool = Body(False, description="是否拼接内容", embed=True),
    is_save_image: bool = Body(False, description="是否保存图片", embed=True),
    use_force_separator: bool = Body(False, description="是否使用强制分隔符", embed=False),
    chunk_type: Optional[str] = Body("tradition", description="切分类型", embed=True),
    sub_chunk_size: Optional[int] = Body(500, description="子块文本块大小", examples=["500"]),
    sub_separator: Optional[list] = Body([], description="子块文本分隔符"),
) -> Response:
    """
    第二步：文件切片预览
    """
    try:
        logger.info(f"开始预览文件切片，知识库ID: {id}, 文件数量: {len(remote_paths)}")
        if (chunk_method == "CharacterTextSplitter" or chunk_method == "SpacyTextSplitter") and (
            separator == "" or separator == []
        ):
            separator = ["\n"]
        elif separator == "":
            separator = []
        from service_knowledge_manage.service.file_processing_service import file_processing_service

        # 这里是直接入库的预览
        success, result = await file_processing_service.preview_file_chunks(
            knowledge_id=id,
            remote_paths=remote_paths,
            chunk_method=chunk_method,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            separator=separator,
            is_generate=is_generate,
            request=request,
            preview_mode=preview_mode,
            multimodal_id=multimodal_id,
            is_header_config=is_header_config,
            start_line=start_line,
            end_line=end_line,
            header_merge_method=header_merge_method,
            is_content_merge=is_content_merge,
            is_save_image=is_save_image,
            use_force_separator=use_force_separator,
            chunk_type=chunk_type,
            sub_chunk_size=sub_chunk_size,
            sub_separator=sub_separator,
        )

        if not success:
            return RetUtil.response_error(message=result)

        # 与file_parse_preview保持一致的返回格式
        if chunk_type != "parent":
            data = {"chunk_type": "tradition", "result": result}
        else:
            data = {"chunk_type": "parent", "result": result}
        return RetUtil.response_ok(data)

    except Exception as e:
        logger.exception("预览文件切片错误", str(e))
        return RetUtil.response_error(message="预览文件切片错误")


@router.post("/file_parse_preview", summary="切片解析")
async def file_parse_preview(
    request: Request,
    id: str = Body(..., description="知识库id", embed=True),
    file_id: str = Body(..., description="文件远程路径列表", embed=True),
    chunk_method: str = Body("RecursiveCharacterTextSplitter", description="切片方式", embed=True),
    chunk_size: int = Body(500, description="文本块大小", embed=True),
    chunk_overlap: int = Body(50, description="文本块重叠大小", embed=True),
    separator: list = Body("", description="文本分隔符", embed=True),
    is_generate: bool = Body(False, description="是否生成问答对", embed=True),
    preview_mode: bool = Body(False, description="文件解析模式(False = 快速模式, True = 解析预览模式)", embed=True),
    multimodal_id: str = Body("", description="多模态模型id", embed=True),
    is_header_config: bool = Body(False, description="是否配置表头", embed=True),
    start_line: int = Body(0, description="开始行", embed=True),
    end_line: int = Body(0, description="结束行", embed=True),
    header_merge_method: str = Body("", description="表头拼接方式", embed=True),
    is_content_merge: bool = Body(False, description="是否拼接内容", embed=True),
    is_save_image: bool = Body(False, description="是否保存图片", embed=True),
    use_force_separator: bool = Body(False, description="是否使用强制分隔符", embed=False),
    chunk_type: Optional[str] = Body(None, description="切分类型", embed=True),
    sub_chunk_size: Optional[int] = Body(None, description="子块文本块大小", examples=["500"]),
    sub_separator: Optional[list] = Body(None, description="子块文本分隔符"),
) -> Response:
    """
    第二步：文件切片预览
    """
    try:
        logger.info(f"开始预览文件切片，知识库ID: {id}, 文件数量: {len(file_id)}")

        if (chunk_method == "CharacterTextSplitter" or chunk_method == "SpacyTextSplitter") and (
            separator == "" or separator == []
        ):
            separator = ["\n"]
        elif separator == "":
            separator = []

        from service_knowledge_manage.service.file_processing_service import file_processing_service

        success, result = await file_processing_service.files_parse_preview(
            file_id=file_id,
            chunk_method=chunk_method,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            separator=separator,
            is_generate=is_generate,
            use_force_separator=use_force_separator,
            chunk_type=chunk_type,
            sub_chunk_size=sub_chunk_size,
            sub_separator=sub_separator,
        )

        if not success:
            return RetUtil.response_error(message=result)
        if chunk_type != "parent":
            data = {"chunk_type": "tradition", "result": result}
        else:
            data = {"chunk_type": "parent", "result": result}
        return RetUtil.response_ok(data)

    except Exception as e:
        logger.exception("预览文件切片错误", str(e))
        return RetUtil.response_error(message="预览文件切片错误")


@router.post("/file_upload_3_v2", summary="向量入库")
async def file_upload_3_v2(
    request: Request,
    id: str = Body(..., description="知识库id", embed=True),
    remote_paths: list[str] = Body(..., embed=True, description="远程文件路径；列表"),
    delete_files: list[str] = Body(..., embed=True, description="要删除的文件ID列表"),
    chunk_method: str = Body("", examples=["RecursiveCharacterTextSplitter"], description="切片方式"),
    chunk_size: int = Body(500, description="文本块大小", examples=["500"]),
    chunk_overlap: int = Body(50, description="文本块重叠大小", examples=["50"]),
    is_generate: bool = Body(False, description="是否生成问答对", examples=[False]),
    separator: list = Body(["\n"], description="文本分隔符"),
    preview_mode: bool = Body(False, description="是否仅预览"),
    multimodal_id: str = Body("", description="多模态模型id", embed=True),
    is_header_config: bool = Body(False, description="是否配置表头", embed=True),
    start_line: int = Body(0, description="开始行", embed=True),
    end_line: int = Body(0, description="结束行", embed=True),
    header_merge_method: str = Body("", description="表头拼接方式", embed=True),
    is_content_merge: bool = Body(False, description="是否拼接内容", embed=True),
    is_save_image: bool = Body(False, description="是否保存图片", embed=True),
    background_tasks: BackgroundTasks = BackgroundTasks(),
    use_force_separator: bool = Body(False, description="是否使用强制分隔符", embed=False),
    chunk_type: Optional[str] = Body(None, description="切分类型", embed=True),
    sub_chunk_size: Optional[int] = Body(None, description="子块文本块大小", examples=["500"]),
    sub_separator: Optional[list] = Body(None, description="子块文本分隔符"),
) -> Response:
    """
    第三步：向量入库
    """
    try:
        logger.info(f"开始快速模式解析并向量入库，知识库ID: {id}, 文件数量: {len(remote_paths)}")

        success, message = await file_processing_service.process_file_upload_batch(
            knowledge_id=id,
            remote_paths=remote_paths,
            delete_files=delete_files,
            chunk_method=chunk_method,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            separator=separator,
            is_generate=is_generate,
            request=request,
            preview_mode=preview_mode,
            multimodal_id=multimodal_id,
            is_header_config=is_header_config,
            start_line=start_line,
            end_line=end_line,
            header_merge_method=header_merge_method,
            is_content_merge=is_content_merge,
            is_save_image=is_save_image,
            background_tasks=background_tasks,
            use_force_separator=use_force_separator,
            chunk_type=chunk_type,
            sub_chunk_size=sub_chunk_size,
            sub_separator=sub_separator,
        )

        # # 启动后台任务处理文件
        # background_tasks.add_task(
        #     _process_file_upload_background,
        #     id,
        #     remote_paths,
        #     delete_files,
        #     chunk_method,
        #     chunk_size,
        #     chunk_overlap,
        #     separator,
        #     is_generate,
        #     request,
        # )
        if preview_mode == False:
            result = await file_processing_service.save_kb_arrange_info(
                id,
                chunk_method,
                chunk_overlap,
                chunk_size,
                is_generate,
                separator,
                use_force_separator,
                chunk_type,
                sub_chunk_size,
                sub_separator,
            )
            if result == False:
                return RetUtil.response_error(message="切片类型与数据库不符!")
        return RetUtil.response_ok("文件已接收，处理中，请稍等")

    except Exception as e:
        logger.exception("上传文件至知识库失败")
        return RetUtil.response_error(message=f"上传文件至知识库失败: {e}")


@router.post("/file_upload_4_v2", summary="向量入库")
async def file_upload_4_v2(
    id: str = Body(..., description="知识库id", embed=True),
    file_id: str = Body(..., description="文件id", embed=True),
    chunk_method: str = Body("", examples=["RecursiveCharacterTextSplitter"], description="切片方式"),
    chunk_size: int = Body(500, description="文本块大小", examples=["500"]),
    chunk_overlap: int = Body(50, description="文本块重叠大小", examples=["50"]),
    is_generate: bool = Body(False, description="是否生成问答对", examples=[False]),
    separator: list = Body(["\n"], description="文本分隔符"),
    use_force_separator: bool = Body(False, description="是否使用强制分隔符", embed=False),
    chunk_type: str = Body(..., description="切分类型", embed=True),
    sub_chunk_size: int = Body(500, description="子块文本块大小", examples=["500"]),
    sub_separator: list = Body(["\n"], description="子块文本分隔符"),
    background_tasks: BackgroundTasks = BackgroundTasks(),
) -> Response:
    """
    第四步：解析预览入库
    """
    try:
        logger.info(f"开始解析预览模式向量入库，文件ID: {file_id}")

        success, message = await file_processing_service.process_file_upload_step_batch(
            file_id=file_id,
            knowledge_id=id,
            chunk_method=chunk_method,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            separator=separator,
            is_generate=is_generate,
            use_force_separator=use_force_separator,
            chunk_type=chunk_type,
            sub_chunk_size=sub_chunk_size,
            sub_separator=sub_separator,
            background_tasks=background_tasks,
        )
        result = await file_processing_service.save_kb_arrange_info(
            id,
            chunk_method,
            chunk_overlap,
            chunk_size,
            is_generate,
            separator,
            use_force_separator,
            chunk_type,
            sub_chunk_size,
            sub_separator,
        )
        if result == False:
            return RetUtil.response_error(message="切片类型与数据库不符!")
        return RetUtil.response_ok("文件已接收，处理中，请稍等")

    except Exception as e:
        logger.exception("上传文件至知识库失败", str(e))
        return RetUtil.response_error(message="上传文件至知识库失败")


@router.post("/get_parse_result_by_file_id", summary="通过文件id获取解析结果")
async def get_parse_result_by_file_id(file_id: str = Body(..., embed=True, description="文件id")):
    """
    文件上传后，前端轮询查询进度
    :param file_id: 文件id
    """
    try:
        if not MongodbUtil.query_doc_by_id(CollectionConfig.UPLOAD_FILE_INFO_COLLECTION, file_id):
            return RetUtil.response_error(message="文件不存在或已删除，无法获取解析结果")
        result = await Knowledge_File_service.get_parse_result_by_file_id(file_id)
        return RetUtil.response_ok(result)
    except Exception as e:
        logger.exception("通过文件id获取解析结果失败", str(e))
        return RetUtil.response_error(message="通过文件id获取解析结果失败")


async def _process_file_upload_background(
    knowledge_id: str,
    remote_paths: list[str],
    delete_files: list[str],
    chunk_method: str,
    chunk_size: int,
    chunk_overlap: int,
    separator: list,
    is_generate: bool,
    request: Request,
):
    """后台处理文件上传任务"""
    try:
        from service_knowledge_manage.service.file_processing_service import file_processing_service

        success, message = await file_processing_service.process_file_upload_batch(
            knowledge_id=knowledge_id,
            remote_paths=remote_paths,
            delete_files=delete_files,
            chunk_method=chunk_method,
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            separator=separator,
            is_generate=is_generate,
            request=request,
        )

        if success:
            logger.info(f"批量文件处理完成: {knowledge_id}")
        else:
            logger.exception(f"批量文件处理失败: {knowledge_id}, 错误: {message}")

    except Exception as e:
        logger.exception(f"后台处理文件上传任务失败: {str(e)}")


@router.post("/get_file_progress", summary="查询文件上传、向量化、入库的进度")
async def get_file_progress(file_id: str = Body(..., embed=True, description="文件id")):
    """
    文件上传后，前端轮询查询进度
    :param file_id: 文件id
    """
    try:
        # 直接调用同步方法获取进度（避免异步序列化问题）
        result = get_progress(file_id)
        logger.info(f"文件上传、向量化、入库的查询进度为：{result}")
        return RetUtil.response_ok(result)
    except Exception as e:
        logger.exception("查询文件进度失败", str(e))
        return RetUtil.response_error(message="查询文件入库进度失败")


@router.post(
    "/chunk_result_query",
    summary="查询文档切片结果",
)
async def chunk_result_query(params: ChunkQueryInfo) -> Response:
    try:

        # logger.info(f"查询知识库文档切片入参信息：{params}")
        result, len_result, chunk_type = await Knowledge_File_service.chunk_result_query_v2(
            params.id, params.file_id, params.page, params.page_size, params.filter_condition
        )
        logger.info(f"查询知识库文档切片数量：{len_result}")
        chunk_list = [chunk["chunk_content"] for chunk in result]

        return RetUtil.response_ok({"total": len_result, "result": result, "chunk_type": chunk_type})
    except Exception as e:
        logger.exception("查询指定知识库中的文件失败", str(e))
        return RetUtil.response_error(message="查询指定知识库中的文件失败")


@router.delete("/file_delete", summary="从指定知识库中删除文件")
async def file_delete(
    id: str = Body(..., description="知识库id", embed=True), file_id: list = Body(..., description="文件id列表")
) -> Response:
    try:
        for item in file_id:
            file_name = MongodbUtil.query_doc_by_id(CollectionConfig.UPLOAD_FILE_INFO_COLLECTION, item).get(
                "file_name", ""
            )

            if file_name:
                MongodbUtil.del_docs_by_condition(
                    CollectionConfig.UPLOAD_FILE_INFO_COLLECTION, del_condition={"knowledge_id": id, "file_id": item}
                )
                await Knowledge_File_service.delete_file(file_id=item, knowledge_id=id, file_name=file_name)
                await asyncio.sleep(1)
        return RetUtil.response_ok("删除成功")
    except Exception as e:
        logger.exception("文件删除失败", str(e))
        return RetUtil.response_error(message="文件删除失败")
    finally:
        for item in file_id:
            MongodbUtil.del_docs_by_condition(
                CollectionConfig.CHUNK_COLLECTION, del_condition={"knowledge_id": id, "file_id": item}
            )


@router.post("/file_query_page", summary="查询指定知识库中的文件(分页)")
async def file_query_page(
    params: FileQueryInfo,
) -> Response:
    try:
        # logger.info(f"查询指定知识库中的文件入参信息：{params}")
        # 1. 查询文件列表（保持原有逻辑）
        result, len_result = await Knowledge_File_service.file_query_page(
            params.id, params.file_name, page=params.page, page_size=params.page_size
        )

        # 2. 为每个文件查询进度并处理时间字段
        for file_item in result:
            file_id = file_item.get("file_id")
            if file_id:
                try:
                    progress_data = get_progress(file_id)
                    if progress_data:
                        # 处理每个阶段的时间字段
                        for stage in progress_data["stages"].values():
                            # 如果存在时间记录则转换为毫秒，否则返回null
                            if stage.get("time_spent") is not None:
                                stage["time_spent"] = int(round(stage["time_spent"] * 1000))
                            else:
                                stage["time_spent"] = None

                        # 处理总时间
                        if progress_data.get("total_time") is not None:
                            progress_data["total_time"] = int(round(progress_data["total_time"] * 1000))
                        else:
                            progress_data["total_time"] = None

                    file_item["progress"] = progress_data
                except Exception as e:
                    logger.exception(f"查询文件进度失败 {file_id}", str(e))
                    file_item["progress"] = None

        # 3. 返回合并后的结果
        logger.info(f"查询到指定知识库中的文件数量为：{len_result}")
        return RetUtil.response_ok({"total": len_result, "result": result})

    except Exception as e:
        logger.exception("查询指定知识库中的文件失败", str(e))
        return RetUtil.response_error(message="查询指定知识库中的文件失败")


@router.post("/file_query_all", summary="查询指定知识库中的所有文件")
async def knowledge_query(
    id: str = Body(..., description="知识库id", embed=True),
    file_name: str = Body("", embed=True, examples=["test_00001"], description="文件名称"),
) -> Response:
    try:
        input_info = {"id": id, "file_name": file_name}
        # logger.info(f"查询指定知识库中的所有文件入参信息:{input_info}")
        result = await Knowledge_File_service.file_query_all(id, file_name)

        def parse_upload_time(upload_time_str):
            return datetime.strptime(upload_time_str, "%Y-%m-%d %H:%M:%S")

        result = sorted(result, key=lambda x: parse_upload_time(x["upload_time"]), reverse=True)
        return RetUtil.response_ok(result)

    except Exception as e:
        logger.exception("查询指定知识库中的文件失败", str(e))
        return RetUtil.response_error(message="查询指定知识库中的文件失败")


@router.post("/update_chunk", summary="保存切片修改")
async def update_chunk(
    index: str = Body(..., embed=True, description="切片下标"),
    new_content: str = Body(..., embed=True, description="新内容"),
    new_question: Optional[str] = Body("", embed=True, description="新内容"),
    id: str = Body(..., description="知识库id", embed=True),
) -> Response:
    try:
        result = None
        for item in MongodbUtil.query_docs_by_condition(
            collection_name=CollectionConfig.CHUNK_COLLECTION, search_condition={"chunk_id": index}
        ):
            result = await Knowledge_File_service.update_chunk_by_parent(new_content, new_question, index)

        if result == None:
            result = await Knowledge_File_service.update_chunk_by_child(int(index), new_content, new_question, id)

        if result:
            logger.info("切片修改成功")
            return RetUtil.response_ok("切片修改成功")
        else:
            return RetUtil.response_ok("切片修改失败")

    except Exception as e:
        logger.exception("保存切片修改失败", str(e))
        return RetUtil.response_error(message="保存切片修改失败")


@router.post("/upload_archive", summary="上传压缩包")
async def upload_archive(
    id: str = Body(..., description="知识库id", embed=True),
    file_obj: UploadFile = File(..., description="压缩包文件"),
) -> Response:
    """
    上传压缩包并解压文件
    :param file_obj: 压缩包文件对象
    :param kb_name: 知识库名称
    :return: 上传结果。
    """
    try:
        input_info = {"id": id}
        # logger.info(f"上传压缩包入参信息:{input_info}")
        results = []
        result = await Knowledge_File_service.upload_and_extract_archive(file_obj, id)
        if len(result["files"]) != 1:
            result["files"] = result["files"][1:]
        for file in result["files"]:
            file_name = os.path.basename(file)
            # 上传文件到MinIO临时桶:tiance-base-temp-file-bucket
            date = datetime.now().strftime("%Y_%m_%d")  # 日期
            folder = generate_unique_id("Temp", datacenter_id=1, worker_id=1)  # 随机id
            remote_path = f"{date}/{folder}/{file_name}"  # 文件路径，以{kb_name}$$${file_name}命名文件名称
            bucket_name = "tiance-base-temp-file-bucket"  # 桶名称
            await run_in_threadpool(MinIoUtil.upload_file, bucket_name, remote_path, file)
            results.append({"file_name": file_name, "remote_path": remote_path})
        logger.info("压缩包上传成功")
        return RetUtil.response_ok(results)

    except Exception as e:
        logger.exception("上传压缩包失败", str(e))
        return RetUtil.response_error(message="上传压缩包失败")

    finally:
        if result:
            for file in result["files"]:
                if os.path.exists(file):
                    os.remove(file)


@router.post("/file_detect", summary="重名/相似文件检测")
async def file_detect(
    id: str = Body(..., examples=["test_upload"], description="知识库id"),
    file_name_list: list = Body(..., examples=[["实验室考勤规则.pdf"]], description="文件名称列表"),
) -> Response:
    try:
        input_info = {"id": id, "file_name_list": file_name_list}
        # logger.info(f"重名/相似文件检测入参信息:{input_info}")
        # 获取重名文件列表
        repeat_file_list = await Knowledge_File_service.repeat_file_detect(id, file_name_list)
        if repeat_file_list == False:
            return RetUtil.response_error(message="获取重名文件失败")

        # 获取相似文件列表
        similar_file_list = await Knowledge_File_service.similar_file_detect(id, file_name_list, repeat_file_list)
        if similar_file_list == False:
            return RetUtil.response_error(message="获取相似文件失败")

        logger.info(f"查询到的重名文件列表为{repeat_file_list}")
        logger.info(f"查询到的相似文件列表为{similar_file_list}")

        return RetUtil.response_ok(
            {
                "repeat_file_list": repeat_file_list,
                "similar_file_list": similar_file_list,
            }
        )

    except Exception as e:
        logger.exception("重名/相似文件检测失败", str(e))
        return RetUtil.response_error(message="重名/相似文件检测失败")


@router.post("/knowledge_list", summary="知识库列表")
async def knowledge_list(
    chat_request: Request,
    team_codes: Optional[list] = Body([], description="团队id", embed=True),
) -> Response:
    try:
        account_id = chat_request.state.account_id
        result = await AgentService.list_knowledge(account_id, team_codes)
        return RetUtil.response_ok(result)
    except Exception as e:
        logger.exception("获取知识库列表失败", str(e))
        return RetUtil.response_error(message="获取知识库列表失败")


@router.post("/get_knowledge_detail", summary="知识库列表")
async def get_knowledge_detail(
    id: str = Body(..., examples=["test_upload"], embed=True),
) -> Response:
    try:
        input_info = {"id": id}
        # logger.info(f"获取知识库详细信息入参信息:{input_info}")
        results = MongodbUtil.query_doc_by_id(collection_name=CollectionConfig.KB_COLLECTION, doc_id=ObjectId(id))
        if not results:
            return RetUtil.response_error(message="知识库不存在")
        for result in [results]:
            result["_id"] = str(result["_id"])
        logger.info(f"知识库详细信息为:{result}")
        return RetUtil.response_ok(result)
    except Exception as e:
        logger.exception("获取知识库详情失败", str(e))
        return RetUtil.response_error(message="获取知识库详情失败")


@router.post("/query_kb_arrange_info", summary="查询知识库编排信息")
async def query_kb_arrange_info(id: str = Body(..., description="知识库id", embed=True)) -> Response:
    """
    第四步：解析预览入库
    """
    try:
        logger.info(f"查询知识库编排信息: {id}")

        result = MongodbUtil.query_doc_by_id(CollectionConfig.KB_ARRANGE_INFO, ObjectId(id))
        if result and result.get("_id", "") != "":
            result["_id"] = str(result["_id"])
            if result["chunk_method"]  in ["parent_by_title","parent_by_paragraph","parent_by_page"]:
                result["chunk_type"] = "parent"
            else:
                result["chunk_type"] = "tradition"

        if result==None:
            # 老数据
            kb_result = MongodbUtil.query_doc_by_id(CollectionConfig.KB_COLLECTION, ObjectId(id))
            if kb_result.get("chunk_type", "") == 1:
                result = {"_id": id}
            else:
                result = {
                    "_id": id,
                    "chunk_method": "CharacterTextSplitter",
                    "chunk_overlap": 0,
                    "chunk_size": 256,
                    "chunk_type": "tradition",
                    "is_generate": False,
                    "separator": [""],
                    "sub_chunk_size": 500,
                    "sub_separator": [""],
                    "use_force_separator": False,
                }
        # result=jsonable_encoder(result)
        return RetUtil.response_ok(result)

    except Exception as e:
        logger.exception("查询知识库编排信息", str(e))
        return RetUtil.response_error(message="查询知识库编排信息")
