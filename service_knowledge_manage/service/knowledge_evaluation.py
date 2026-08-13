from sqlalchemy import or_, func, distinct, exists, and_, desc, text, case
from sqlalchemy.exc import SQLAlchemyError
from celery.result import AsyncResult
from service_celery_manage.celery_app import celery_app
from fileinput import filename
from service_knowledge_manage.entity.knowledge_hub_entity import KnowledgeRetrivalInfo
from service_knowledge_manage.service.knowledge_retrieval_service import knowledge_retrieval_service
from service_knowledge_manage.entity.knowledge_entity import EvaluationQuestion
import pandas as pd
from service_knowledge_manage.entity.knowledge_entity import KnowledgeEvaluationRequest
import asyncio
import os
import time
from pathlib import Path
from typing import Any, Optional, Dict, List

import numpy as np
import pandas as pd
from bson import ObjectId
from fastapi import BackgroundTasks, Depends, Request
from fastapi.concurrency import run_in_threadpool
from loguru import logger
from requests import Session
from sqlalchemy.orm import Session

from base_configs.api_config import ApiConfig
from base_configs.mongodb_config import CollectionConfig
from base_utils.embedding_util import EmbeddingUtil
from base_utils.excel_util import ExcelUtil
from base_utils.minio_util import MinIoUtil
from base_utils.mongodb_util import MongodbUtil
from sqlalchemy.orm import Session
from base_configs.model_config import ModelConfig
from base_utils.embedding_util import EmbeddingUtil
from service_knowledge_manage.entity.knowledge_entity import KNOWLEDGE_EVALUATION, KnowledgeEvaluationSetting
from service_knowledge_manage.entity.knowledge_entity import EvaluationResult, EvaluationAnswer
from base_configs.mongodb_config import CollectionConfig
from bson import ObjectId
from sklearn.metrics.pairwise import cosine_similarity as pairwise_cosine_similarity
# 移除对 torch/torchmetrics 的依赖，改为纯 Python 计算 MRR 与 Recall@K
from fastapi.concurrency import run_in_threadpool
from base_utils.minio_util import MinIoUtil
from fastapi import Request, BackgroundTasks
from typing import Optional
# class Knowledge_Evaluation:
# '''
# Author: Sandra_feng
# Date: 2025-10-31 17:21:42
# FilePath: \tiance-base\service_knowledge_manage\service\knowledge_evaluation.py
# Description: 知识库评估服务
# '''
from service_knowledge_manage.entity.knowledge_entity import RetrivalInfo
from loguru import logger
import os
import uuid
import math
from datetime import datetime, timedelta
from pathlib import Path

from service_celery_manage.evaluation_tasks import kb_evaluation_task
from service_celery_manage.tasks import file_parse_task, split_embedding_task
from service_knowledge_manage.entity.knowledge_entity import KNOWLEDGE_EVALUATION, RetrivalInfo
from service_knowledge_manage.entity.knowledge_hub_entity import KnowledgeRetrivalInfo
from service_knowledge_manage.service.file_processing_service import FileProcessingService
from service_knowledge_manage.service.knowledge_file_service import Knowledge_File_service
from service_knowledge_manage.service.parse_service import FileParseService
from service_knowledge_manage.service.splitter_service import SplitterService
from service_knowledge_manage.service.util.file_progress import set_progress

def get_db(request: Request):
    db = request.app.state.SessionLocal()
    try:
        yield db
    finally:
        db.close()
'''
Author: Sandra_feng
Date: 2025-10-31 17:21:42
FilePath: \tiance-base\service_knowledge_manage\service\knowledge_evaluation.py
Description: 知识库评估服务
'''

class Knowledge_Evaluation_service:
    """
    知识库评估服务
    """
    def _load_stopwords(self) -> set:
        try:
            if hasattr(self, "_stopwords") and isinstance(self._stopwords, set) and len(self._stopwords) > 0:
                return self._stopwords
        except Exception:
            pass
        from pathlib import Path
        path = str(Path(__file__).parent / "util" / "StopWords" / "baidu_stopwords.txt")
        s = set()
        try:
            with open(path, "r", encoding="utf-8") as f:
                for line in f:
                    w = line.strip()
                    if w:
                        s.add(w)
            # logger.info(f"已加载停用词表，路径：{path}")
        except Exception:
            logger.error(f"加载停用词表失败，路径：{path}")
            s = set()
        self._stopwords = s
        return self._stopwords

    def _tokenize(self, text: str) -> list[str]:
        import re
        sw = self._load_stopwords()
        tokens: list[str] = []
        try:
            import jieba  # type: ignore
            use_jieba = True
        except Exception:
            use_jieba = False
        if not use_jieba:
            try:
                import pkuseg  # type: ignore
                if not hasattr(self, "_pkuseg"):
                    self._pkuseg = pkuseg.pkuseg()
            except Exception:
                self._pkuseg = None
        try:
            import wordninja  # type: ignore
        except Exception:
            wordninja = None
        parts = re.findall(r"[\u4e00-\u9fff]+|[A-Za-z0-9]+", text)
        for p in parts:
            if re.match(r"^[\u4e00-\u9fff]+$", p):
                if use_jieba:
                    segs = list(jieba.lcut(p))  # type: ignore
                elif getattr(self, "_pkuseg", None) is not None:
                    segs = list(self._pkuseg.cut(p))
                else:
                    segs = [p]
            else:
                if wordninja is not None and re.match(r"^[A-Za-z]+$", p):
                    segs = wordninja.split(p.lower())  # type: ignore
                else:
                    segs = [p.lower()]
            for t in segs:
                if not t:
                    continue
                if t in sw:
                    continue
                if re.match(r"^[\u4e00-\u9fffA-Za-z0-9]+$", t):
                    tokens.append(t)
        return tokens


    async def save_evaluation_setting(self, db: Session, request: KnowledgeEvaluationRequest, file_id: str, *, status: int = 1) -> bool:
        """
        将评测相关的检索配置全部写入 MySQL
        """
        try:
            params = getattr(request, "retrival_params", None)
            # 计算融合权重（语义/关键词）
            fusion_weights = []
            if params and hasattr(params, "semantics_weights") and hasattr(params, "keywords_weights"):
                fusion_weights = [params.semantics_weights, params.keywords_weights]
            else:
                fusion_weights = [0.7, 0.3]

            # 使用传入的 evaluation_id 或 request.task_id
            eval_id=request.evaluation_id
            if not eval_id:
                raise ValueError("evaluation_id 不能为空，必须提供 evaluation_id 或 request.task_id")
            rerank_id = getattr(params, "rerank_id", "")

            if rerank_id != '':
                model_data = MongodbUtil.query_doc_by_id(
                    collection_name=CollectionConfig.MODEL_RUN_COLLECTION,
                    doc_id=ObjectId(rerank_id),
                )
                if model_data:
                    rerank_model = model_data.get("model_uid", "")
            else:
                rerank_model = getattr(params, "rerank_model", "")

            # 映射到 KnowledgeEvaluationSetting 的字段进行保存
            setting = KnowledgeEvaluationSetting(
                evaluation_id=eval_id,
                file_id=file_id,
                evaluation_name=request.task_name,
                similarity_threshold=request.similarity_threshold,
                search_type=getattr(params, "search_type", "semantic"),
                recall_num=getattr(params, "recall_num", 0),
                is_rerank=bool(getattr(params, "is_rerank", False)),
                rerank_id=rerank_id,
                rerank_model=rerank_model ,
                rerank_num=getattr(params, "rerank_num", 0),
                score=getattr(params, "score", 0.0),
                needEnhanceRounds=bool(getattr(params, "enhance_rounds", 0)),
                enhance_rounds=getattr(params, "enhance_rounds", 0),
                semantics_weights=fusion_weights[0],
                keywords_weights=fusion_weights[1],
                evaluation_num=0,
                create_time=datetime.now().strftime("%Y%m%d%H%M%S"),
                update_time=datetime.now().strftime("%Y%m%d%H%M%S"),
                status=status,
                message=""
            )

            db.add(setting)
            db.commit()
            logger.info(f"评测检索配置已保存到 MySQL（KnowledgeEvaluationSetting），file_id={file_id}")
            return True
        except Exception as e:
            logger.exception(f"保存评测检索配置失败: {e}")
            raise
    async def process_evaluation(self, file_path: str, file_id: str, request: KnowledgeEvaluationRequest, db: Session) -> dict:
        """
        处理知识库评估任务
        """
        try:

            # 更新评测状态至 2（评测中）
            evaluation_id = request.evaluation_id
            db.query(KnowledgeEvaluationSetting).filter(
                KnowledgeEvaluationSetting.evaluation_id == evaluation_id
            ).update({
                KnowledgeEvaluationSetting.status: 2,
                KnowledgeEvaluationSetting.update_time: datetime.now().strftime("%Y%m%d%H%M%S"),
                KnowledgeEvaluationSetting.message: None
            })
            db.commit()
            # 1. 根据 file_id 查询问题与标准答案，并按创建时间排序
            question_rows = db.query(EvaluationQuestion). \
                filter(EvaluationQuestion.file_id == file_id). \
                order_by(EvaluationQuestion.question_index.asc(), EvaluationQuestion.create_time.asc()).all()
            # questions = [q.question for q in question_rows]
            # standard_answers = [q.standard_answer for q in question_rows]
            # 2. 初始化评测数据（evaluation_id 已在上面确定）
            evaluation_result = []
            answers_data = []
            qa_map: dict[str, dict] = {}
            total_hit_num = 0
            total_not_hit_num = 0
            # 3. 遍历问题进行检索和相似度计算
            for q in question_rows:
                NOT_HIT_FLAG=True
                question_id = q.question_id
                question = q.question
                standard_answer = q.standard_answer
                qa_map[question_id] = {"question": question, "standard_answer": standard_answer}
                # 3.2 调用检索服务进行检索
                rerank_id=request.retrival_params.rerank_id
                if rerank_id != '':
                    model_data = MongodbUtil.query_doc_by_id(
                        collection_name=CollectionConfig.MODEL_RUN_COLLECTION,
                        doc_id=ObjectId(rerank_id),
                    )
                    if model_data:
                        rerank_model = model_data.get("model_uid", "")
                else:
                    rerank_model = request.retrival_params.rerank_model
                retrival_params = KnowledgeRetrivalInfo(
                    id=request.knowledge_id,
                    user_query=question,
                    search_type=request.retrival_params.search_type,
                    recall_num=request.retrival_params.recall_num,
                    rerank_id=rerank_id,
                    rerank_model=rerank_model,
                    rerank_num=request.retrival_params.rerank_num,
                    score=request.retrival_params.score,
                    enhance_rounds=request.retrival_params.enhance_rounds,
                    filter=request.retrival_params.filter,
                    fusion_weights=[request.retrival_params.semantics_weights, request.retrival_params.keywords_weights],
                )
                retrival_result = await knowledge_retrieval_service.advanced_knowledge_retrieval(retrival_params)
                result_dict = retrival_result.model_dump()
                if result_dict["results"]:
                    k = len(result_dict["results"])
                    # logger.info(f"检索结果数量: {k}")
                else:#检索结果为空
                    NOT_HIT_FLAG = False
                    k = request.retrival_params.recall_num
                    total_not_hit_num += 1
                    answer_id = uuid.uuid4().hex
                    ans = {
                        "answer_id": answer_id,
                        "question_id": question_id,
                        "evaluation_id": evaluation_id,
                        "chunk_content": "",
                        "recall_score": 0.0,
                        "is_hit": False,
                        "similarity": 0.0,
                        "index": -1,
                        "hit_score": 0.0,
                        "all_hit": False,
                    }
                    answers_data.append(ans)
                    db.add(EvaluationAnswer(**ans))
                # 3.3 计算检索答案与标准答案的相似度和是否命中
                docs_tokens = []
                for _c in result_dict["results"]:
                    _text = (_c["chunk_content"] or "").replace("\n", "")
                    docs_tokens.append(self._tokenize(_text))
                
                query_tokens = self._tokenize(standard_answer or "")
                # logger.info(f"标准答案的词集: {query_tokens}")
                all_doc_tokens = set()
                for _t in docs_tokens:
                    for _x in _t:
                        all_doc_tokens.add(_x)
                _ql = len(query_tokens) ##标准答案长度
                if _ql == 0:
                    group_similarity_score = 0.0
                else:
                    matched_tokens = [t for t in query_tokens if t in all_doc_tokens]
                    _mc = len(matched_tokens)
                    group_similarity_score = math.floor((_mc / float(_ql)) * 100.0) / 100.0
                logger.info(f"相匹配的词集: {matched_tokens}")
                logger.info(f"标准答案的词集长度✨: {_ql}")
                logger.info(f"所有检索切片与标准答案命中相似度✨: {group_similarity_score}")
                be_hited = False
                if group_similarity_score >= 0.8: ##标准答案与检索切片相似度大于0.8则认为命中
                    be_hited = True
                    total_hit_num += 1
                elif NOT_HIT_FLAG:
                    total_not_hit_num += 1

                for idx, chunk in enumerate(result_dict["results"]):
                    answer_id = uuid.uuid4().hex
                    retrival_answer = chunk["chunk_content"].replace("\n", "")
                    doc_tokens = self._tokenize(retrival_answer)
                    # logger.info(f"检索切片{idx}的词集: {doc_tokens}")
                    
                    q_len = len(query_tokens)
                    if q_len == 0:
                        similarity_score = 0.0
                    else:
                        doc_set = set(doc_tokens)
                        match_cnt = sum(1 for t in query_tokens if t in doc_set)
                        similarity_score = math.floor((match_cnt / float(q_len)) * 100.0) / 100.0
                    # logger.info(f"匹配相似度: {similarity_score}")

                    ans = {
                        "answer_id": answer_id,
                        "question_id": question_id,
                        "evaluation_id": evaluation_id,
                        "chunk_content": chunk["chunk_content"],
                        "recall_score": chunk["recall_score"],
                        "is_hit": similarity_score >= request.similarity_threshold,
                        "similarity": similarity_score,
                        "index": chunk["recall_index"],
                        "hit_score":group_similarity_score,
                        "all_hit": be_hited,
                    }
                    answers_data.append(ans)
                    db.add(EvaluationAnswer(**ans))

            db.commit()
            # 4.2/4.3 计算 MRR 和 Recall@K（封装为函数）
            # k = len(chunk_contents)
            mrr_value, recall_k = self.compute_metrics(answers_data, top_k=k if (k and k > 0) else None)

            # 4.4 保存评测数据
            evaluation_result.append({
                "evaluation_result_id": uuid.uuid4().hex,
                "evaluation_id": evaluation_id,
                "search_type": request.retrival_params.search_type,
                "update_time": datetime.now().strftime("%Y%m%d%H%M%S"),
                "create_time": datetime.now().strftime("%Y%m%d%H%M%S"),
                "recall_k":recall_k,
                "mrr":mrr_value,
                "hit_num": total_hit_num,# 汇总所有问题的命中/未命中总数
                "not_hit_num": total_not_hit_num,
            })
            # 4.5 保存评测结果到 MySQL（evaluation_result表）
            evaluation_result = EvaluationResult(**evaluation_result[0])
            db.add(evaluation_result)
            db.commit()
            # 评测完成，更新状态为 3
            db.query(KnowledgeEvaluationSetting).filter(
                KnowledgeEvaluationSetting.evaluation_id == evaluation_id
            ).update({
                KnowledgeEvaluationSetting.status: 3,
                KnowledgeEvaluationSetting.update_time: datetime.now().strftime("%Y%m%d%H%M%S"),
            })
            db.commit()
            logger.info("process_evaluation 业务逻辑已全部完成😊，即将 return")
            return {
                "success": True,
                "message": "评测成功",
                # "excel_path": evaluation_file_path,
            }
        except Exception as e:
            logger.exception(f"评测流程失败: {e}")
            # 评测失败，更新状态为 4，并记录错误信息
            try:
                db.query(KnowledgeEvaluationSetting).filter(
                    KnowledgeEvaluationSetting.evaluation_id == request.evaluation_id
                ).update({
                    KnowledgeEvaluationSetting.status: 4,
                    KnowledgeEvaluationSetting.message: str(e),
                    KnowledgeEvaluationSetting.update_time: datetime.now().strftime("%Y%m%d%H%M%S"),
                })
                db.commit()
            except Exception as _:
                db.rollback()
            return {
                "success": False,
                "message": f"评测失败: {e}",
            }
        
    @staticmethod
    def compare_task_metrics(task1: EvaluationResult, task2: EvaluationResult) -> dict:
        """
        对比两个评测任务的指标，重点计算：新增命中数、以及新增未命中数。
        返回：
        - dict，包含两个任务的核心指标，以及增量（task2 相比 task1 的变化量）
        """
        # 安全取数，避免 None
        t1_hit = int(task1.hit_num or 0)
        t1_not_hit = int(task1.not_hit_num or 0)
        t2_hit = int(task2.hit_num or 0)
        t2_not_hit = int(task2.not_hit_num or 0)

        # 度量差值，保留两位小数。注意双精度下无法精确表示，做减法时只能得到近似值
        _t1_recall_i = math.floor(float(getattr(task1, "recall_k", 0.0) or 0.0) * 100.0)
        _t2_recall_i = math.floor(float(getattr(task2, "recall_k", 0.0) or 0.0) * 100.0)
        _t1_mrr_i = math.floor(float(getattr(task1, "mrr", 0.0) or 0.0) * 100.0)
        _t2_mrr_i = math.floor(float(getattr(task2, "mrr", 0.0) or 0.0) * 100.0)
        t1_recall_k = _t1_recall_i / 100.0
        t2_recall_k = _t2_recall_i / 100.0
        t1_mrr = _t1_mrr_i / 100.0
        t2_mrr = _t2_mrr_i / 100.0
        logger.info(f"💐t1_recall_k: {t1_recall_k}, t2_recall_k: {t2_recall_k}, t1_mrr: {t1_mrr}, t2_mrr: {t2_mrr}")

        comparison = {
            "hit_diff": (t2_hit - t1_hit) if (t2_hit - t1_hit) >= 0 else 0,
            "not_hit_diff": (t2_not_hit - t1_not_hit) if (t2_not_hit - t1_not_hit) >= 0 else 0,
            "recallk_diff": (_t2_recall_i - _t1_recall_i) / 100.0,
            "mrr_diff": (_t2_mrr_i - _t1_mrr_i) / 100.0,
        }
        return comparison
    
    def compute_metrics(self, answers_data: list[dict], top_k: Optional[int] = None) -> tuple[float, float]:
        """
        使用纯 Python 计算检索评测指标：适应不同的环境，不依赖torch和tf环境
        - MRR（Mean Reciprocal Rank）：每个问题的第一个命中条目的倒数排名的平均值。
        - Recall@K（或 HitRate@K）：在前 K 条内有命中的问题占比。

        参数：
        - answer_data: 由检索结果组成的列表，每个元素至少包含 question_id、index（排名）、is_hit 字段。
        - top_k: Recall 的 K 值；不传则自动按每个问题的最大检索条数推断一个全局 K。

        返回：
        - (mrr_value, recall_k)
        """
        if not answers_data:
            return 0.0, 0.0

        # 判断是否为 0-based 排名，并按问题分组后按 index 排序
        zero_based = min(d.get("index", 0) for d in answers_data) == 0
        groups: dict[str, list[dict]] = {}
        for d in answers_data:
            qid = d.get("question_id")
            if qid is None:
                # 跳过缺失 question_id 的数据
                continue
            groups.setdefault(qid, []).append(d)

        # 为每个问题按 index 排序，确保排名顺序正确
        for qid in groups:
            groups[qid].sort(key=lambda x: x.get("index", 0))

        num_queries = len(groups)
        if num_queries == 0:
            return 0.0, 0.0

        mrr_sum = 0.0
        for items in groups.values():
            rr = 0.0
            for item in items:
                if item.get("is_hit"):
                    rank = item.get("index", 0) + 1 if zero_based else item.get("index", 1)
                    if rank <= 0:
                        rank = 1
                    rr = 1.0 / float(rank)
                    break
            mrr_sum += rr
        mrr_value = mrr_sum / num_queries

        hits = 0
        for items in groups.values():
            if any(i.get("all_hit") for i in items):
                hits += 1
        recall_k = hits / num_queries

        return math.floor(mrr_value * 100.0) / 100.0, math.floor(recall_k * 100.0) / 100.0

    @staticmethod
    def save_to_excel(evaluation_result: dict,
                      file_path: str,
                      file_id: str,
                      sheet_name: str = "Evaluation Results",
                      answers: list[dict] | None = None,
                      qa_map: dict[str, dict] | None = None,
                      hit_threshold: float = 0.0) -> str:
        """
        将评测结果按图示样式写入 Excel 并返回生成的文件路径。
        """
        # 1) 生成保存路径
        src_path = Path(file_path)
        base_dir = src_path.parent
        # base_name = src_path.stem
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        excel_name = f"{file_id}_{timestamp}_results.xlsx"
        excel_path = base_dir / excel_name

        # 2) 使用 openpyxl 按分区样式写入
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        base_dir.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 样式
        header_fill = PatternFill("solid", fgColor="4F81BD")
        header_font = Font(color="FFFFFF", bold=True)
        label_align = Alignment(horizontal="left", vertical="center")
        value_align = Alignment(horizontal="left", vertical="center")
        thin = Side(style="thin", color="999999")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 60

        def set_row(r, label, value=None, is_header=False):
            ws.cell(r, 1, label)
            ws.cell(r, 2, value if value is not None else "")
            # 样式
            ws.cell(r, 1).alignment = label_align
            ws.cell(r, 2).alignment = value_align
            ws.cell(r, 1).border = border
            ws.cell(r, 2).border = border
            if is_header:
                ws.cell(r, 1).fill = header_fill
                ws.cell(r, 2).fill = header_fill
                ws.cell(r, 1).font = header_font
                ws.cell(r, 2).font = header_font

        # 数据映射
        def map_search_type(st: str) -> str:
            mapping = {
                "semantic": "向量检索",
                "fulltext": "全文检索",
                "hybrid": "混合检索",
            }
            return mapping.get(str(st).lower(), str(st))

        row = 1
        # 测评任务 & 时间
        set_row(row, "测评任务", evaluation_result.get("task_name", ""), is_header=True); row += 1
        set_row(row, "测评时间", evaluation_result.get("evaluation_time", "")); row += 1
        logger.info(f"测评任务 & 时间: {evaluation_result.get('task_name', '')} - {evaluation_result.get('evaluation_time', '')}")

        # 测评参数
        # rank_model = evaluation_result.get("rank_model", "")
        # logger.info(f"重排模型💐💐💐：{rank_model}")
        set_row(row, "测评参数", "", is_header=True); row += 1
        params = [
        ("检索方式", map_search_type(evaluation_result.get("search_type", "semantic"))),
        ("检索条数", evaluation_result.get("recall_num", "")),
        ("重排模型", evaluation_result.get("rerank_model", "")),
        ("重排条数", evaluation_result.get("rerank_num", "")),
        ("重排相似度阈值", evaluation_result.get("rerank_threshold", "")) if evaluation_result.get("rerank_model") else ("重排相似度阈值", ""),
        ("增强条数", evaluation_result.get("enhance_rounds", "")),
        ("混合检索权重", evaluation_result.get("fusion_weights", "")) if evaluation_result.get("search_type") == "hybrid" else ("混合检索权重", ""),
        ("命中相似度阈值", hit_threshold),
        ]
        for label, val in params:
            set_row(row, label, val); row += 1

        # 测评结果
        set_row(row, "测评结果", "", is_header=True); row += 1
        results = [
            ("Recall@K", evaluation_result.get("recall_k", "")),
            ("MRR", evaluation_result.get("mrr", "")),
            ("命中数", evaluation_result.get("hit_num", "")),
            ("未命中数", evaluation_result.get("not_hit_num", "")),
        ]
        for label, val in results:
            set_row(row, label, val); row += 1

        # 回答详情sheet2
        sheet2 = wb.create_sheet(title="测评详情")
        table_header_fill = PatternFill("solid", fgColor="4F81BD")
        table_header_font = Font(color="FFFFFF", bold=True)
        normal_font = Font(color="000000")
        red_font = Font(color="CC0000")
        bold_font = Font(color="000000", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

        def set_sheet2_row(r, c, value, *, header=False, align="left", font=None):
            cell = sheet2.cell(r, c, value)
            cell.border = border
            if header:
                cell.fill = table_header_fill
                cell.font = table_header_font
                cell.alignment = center_align
            else:
                cell.font = font or normal_font
                cell.alignment = center_align if align == "center" else left_align
            return cell

        def _auto_fit_row_height(r: int, start_col: int, end_col: int, text: str):
            from math import ceil
            from openpyxl.utils import get_column_letter
            import unicodedata
            if text is None:
                text = ""
            total_width = 0.0
            for c in range(start_col, end_col + 1):
                letter = get_column_letter(c)
                w = sheet2.column_dimensions[letter].width or 8.38
                total_width += float(w)
            per_line = max(1, int(total_width))
            parts = str(text).split("\n")
            lines = 0
            for p in parts:
                eff_len = 0
                for ch in p:
                    ea = unicodedata.east_asian_width(ch)
                    eff_len += 2 if ea in ("W", "F") else 1
                lines += ceil(max(1, eff_len) / per_line)
            size = sheet2.cell(r, 2).font.sz or 11
            base = float(size) * 1.6
            sheet2.row_dimensions[r].height = max(base, lines * base)

        def _auto_fit_columns():
            from openpyxl.utils import get_column_letter
            import unicodedata
            max_len = {1: 6, 2: 6, 3: 6, 4: 6, 5: 6, 6: 6, 7: 6}
            for row in sheet2.iter_rows(values_only=True):
                for idx in range(7):
                    v = row[idx] if idx < len(row) else ""
                    s = str(v) if v is not None else ""
                    l = 0
                    for ch in s:
                        ea = unicodedata.east_asian_width(ch)
                        l += 2 if ea in ("W", "F") else 1
                    if l > max_len[idx + 1]:
                        max_len[idx + 1] = l
            for i in range(1, 8):
                letter = get_column_letter(i)
                sheet2.column_dimensions[letter].width = min(100, max(8.38, max_len[i] + 2))

        cur = 1
        headers = ["问题", "标准答案", "命中分数", "是否命中", "切片内容", "召回分数", "相似度"]
        for col_idx, h in enumerate(headers, start=1):
            set_sheet2_row(cur, col_idx, h, header=True)
        cur += 1

        if answers:
            grouped: dict[str, list[dict]] = {}
            for row in answers:
                grouped.setdefault(row["question_id"], []).append(row)
            ordered_qids = sorted(grouped.keys(), key=lambda qid: int(((qa_map or {}).get(qid, {}).get("question_index", 0) or 0)))
            for qid in ordered_qids:
                rows = grouped.get(qid, [])
                rows.sort(key=lambda x: x.get("index", 0))
                q_info = (qa_map or {}).get(qid, {"question": "", "standard_answer": ""})
                hit_flag = any(bool(x.get("all_hit")) for x in rows)
                hit_score_val = float(rows[0].get("hit_score", 0.0)) if rows else 0.0
                start_r = cur
                for d in rows:
                    set_sheet2_row(cur, 5, d.get("chunk_content", ""), align="left")
                    set_sheet2_row(cur, 6, round(float(d.get("recall_score", 0.0)), 2), align="center")
                    sim = float(d.get("similarity", 0.0))
                    set_sheet2_row(cur, 7, round(sim, 2), align="center", font=(red_font if sim < float(hit_threshold) else normal_font))
                    cur += 1
                if cur == start_r:
                    set_sheet2_row(cur, 5, "", align="left"); set_sheet2_row(cur, 6, "", align="center"); set_sheet2_row(cur, 7, "", align="center"); cur += 1
                end_r = cur - 1
                set_sheet2_row(start_r, 1, q_info.get("question", ""), align="center")
                set_sheet2_row(start_r, 2, q_info.get("standard_answer", ""), align="center")
                set_sheet2_row(start_r, 3, round(hit_score_val, 2), align="center", font=(red_font if hit_score_val < 0.8 else normal_font))
                set_sheet2_row(start_r, 4, "是" if hit_flag else "否", align="center", font=(bold_font if hit_flag else red_font))
                sheet2.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=1)
                sheet2.merge_cells(start_row=start_r, start_column=2, end_row=end_r, end_column=2)
                sheet2.merge_cells(start_row=start_r, start_column=3, end_row=end_r, end_column=3)
                sheet2.merge_cells(start_row=start_r, start_column=4, end_row=end_r, end_column=4)
                for rr in range(start_r, end_r + 1):
                    sheet2.cell(rr, 1).border = border; sheet2.cell(rr, 1).alignment = center_align
                    sheet2.cell(rr, 2).border = border; sheet2.cell(rr, 2).alignment = center_align
                    sheet2.cell(rr, 3).border = border; sheet2.cell(rr, 3).alignment = center_align
                    sheet2.cell(rr, 4).border = border; sheet2.cell(rr, 4).alignment = center_align
                _auto_fit_row_height(start_r, 1, 2, q_info.get("standard_answer", ""))

        _auto_fit_columns()

        wb.save(excel_path)
        return str(excel_path)

    @staticmethod
    def batch_delete_evaluation_data(db: Session, file_ids: List[str]) -> dict:
        """
        批量删除文件ID列表对应的所有测评关联数据
        逻辑：
        1. 遍历每个file_id，依次删除其关联的所有表数据
        2. 表删除顺序：问题表→答案表→结果表→设置表→主表（保证外键约束不冲突）
        3. 统计每个文件的删除条数及总条数
        """
        try:
            total_delete_count = 0

            for file_id in file_ids:
                file_delete_count = 0  # 单个文件的删除条数统计
                setting = db.query(KnowledgeEvaluationSetting).filter(
                    KnowledgeEvaluationSetting.file_id == file_id,
                    KnowledgeEvaluationSetting.status.in_([1, 2])
                ).order_by(KnowledgeEvaluationSetting.update_time.desc()).first()
                if setting:
                    Knowledge_Evaluation_service.abort_evaluation_task(db=db, file_id=file_id)

                eq_delete = db.query(EvaluationQuestion).filter(
                    EvaluationQuestion.file_id == file_id
                ).delete(synchronize_session=False)
                file_delete_count += eq_delete

                settings = db.query(KnowledgeEvaluationSetting.evaluation_id).filter(
                    KnowledgeEvaluationSetting.file_id == file_id
                ).all()
                evaluation_ids = [item.evaluation_id for item in settings] if settings else []

                ea_delete = 0
                if evaluation_ids:
                    ea_delete = db.query(EvaluationAnswer).filter(
                        EvaluationAnswer.evaluation_id.in_(evaluation_ids)
                    ).delete(synchronize_session=False)
                file_delete_count += ea_delete

                er_delete = 0
                if evaluation_ids:
                    er_delete = db.query(EvaluationResult).filter(
                        EvaluationResult.evaluation_id.in_(evaluation_ids)
                    ).delete(synchronize_session=False)
                file_delete_count += er_delete

                es_delete = 0
                if evaluation_ids:
                    es_delete = db.query(KnowledgeEvaluationSetting).filter(
                        KnowledgeEvaluationSetting.evaluation_id.in_(evaluation_ids)
                    ).delete(synchronize_session=False)
                file_delete_count += es_delete

                ke_delete = db.query(KNOWLEDGE_EVALUATION).filter(
                    KNOWLEDGE_EVALUATION.file_id == file_id
                ).delete(synchronize_session=False)
                file_delete_count += ke_delete


                total_delete_count += file_delete_count
            # 提交事务（所有文件的删除操作在一个事务中，确保原子性）
            db.commit()
            return {
                "success": True,
                "message": f"批量删除完成",
                "delete_count": total_delete_count,
            }
        except Exception as e:
            # 异常时回滚所有操作
            db.rollback()
            logger.exception(f"批量删除测评数据失败：{str(e)}")
            raise e


    def get_hit_stat(
            db: Session,
            file_id: str,
    ):
        """通过文件ID查询所有有效测评的统计结果：
        1. 查询该文件下所有状态3/4/5的测评记录（按创建时间倒序）
        2. 逐个获取每条测评的统计结果，汇总返回
        """
        # 2. 逐个处理每条测评的统计结果
        total_question_all = 0  # 所有测评总问题数
        hit_count_all = 0  # 所有测评总命中数
        no_hit_count_all = 0  # 所有测评总未命中数
        evaluation_list = []  # 单条测评详情列表
        # 1. 查询该文件下所有有效测评记录（状态3/4/5，按创建时间倒序）
        all_settings = db.query(KnowledgeEvaluationSetting
        ).filter(
            KnowledgeEvaluationSetting.file_id == file_id,
            KnowledgeEvaluationSetting.status.in_([3, 4, 5])
        ).order_by(
            desc(KnowledgeEvaluationSetting.create_time)
        ).all()

        if not all_settings:
            return {
                "total_statistics": {
                    "total_question_all": total_question_all,
                    "hit_count_all": hit_count_all,
                    "no_hit_count_all": no_hit_count_all,
                    "hit_rate": round(hit_count_all / total_question_all, 4) if total_question_all > 0 else 0.0
                },
                "evaluation_details": evaluation_list  # 无数据时为[]
            }

        for setting in all_settings:
            evaluation_id = setting.evaluation_id
            # 查询当前测评的统计结果
            eval_result = db.query(
                EvaluationResult.hit_num,
                EvaluationResult.not_hit_num,
                EvaluationResult.recall_k,
                EvaluationResult.mrr,
                EvaluationResult.evaluation_result_url,
                EvaluationResult.search_type,
            ).filter(EvaluationResult.evaluation_id == evaluation_id).first()

            # 统计字段默认值处理
            hit_num = eval_result.hit_num or 0 if eval_result else 0
            not_hit_num = eval_result.not_hit_num or 0 if eval_result else 0
            total_question = hit_num + not_hit_num
            # 累加总统计
            total_question_all += total_question
            hit_count_all += hit_num
            no_hit_count_all += not_hit_num

            # 组装单条测评结果
            evaluation_list.append({
                "file_id": setting.file_id,
                "evaluation_id": evaluation_id,
                "evaluation_name": setting.evaluation_name or "未命名评测",
                "status": setting.status,
                "message": setting.message or "",
                "evaluation_create_time": setting.create_time.strftime("%Y-%m-%d %H:%M:%S") if setting.create_time else "",
                "hit_num": hit_num,
                "no_hit_num": not_hit_num,
                # 保留合法的 0.0，不要因为真假判断把 0.0 变成 None
                "recall_k": float(eval_result.recall_k) if (eval_result is not None and eval_result.recall_k is not None) else 0.0,
                "mrr": float(eval_result.mrr) if (eval_result is not None and eval_result.mrr is not None) else 0.0,
                "evaluation_result_url": eval_result.evaluation_result_url if eval_result else "",
                "similarity_threshold": setting.similarity_threshold,
                "create_time": setting.create_time.strftime("%Y-%m-%d %H:%M:%S") if setting.create_time else None,
                "update_time": setting.update_time.strftime("%Y-%m-%d %H:%M:%S") if setting.update_time else None,
                "search_type": setting.search_type,
                "recall_num": setting.recall_num,
                "is_rerank": setting.is_rerank,
                "rerank_id": setting.rerank_id,
                "rerank_model": setting.rerank_model,
                "rerank_num": setting.rerank_num,
                "score": setting.score,
                "needEnhanceRounds": setting.needEnhanceRounds,
                "enhance_rounds": setting.enhance_rounds,
                "semantics_weights": setting.semantics_weights,
                "keywords_weights": setting.keywords_weights,
                "evaluation_num": setting.evaluation_num,
            })

        # 3. 组装最终返回结果（包含总体统计 + 单条详情）
        return evaluation_list

    @staticmethod
    def get_evaluation_dataset(
            db: Session,
            knowledge_id: str,
            file_name: Optional[str] = None,
            page: int = 1,
            page_size: int = 10
    ) -> dict:
        """
        知识库下所有评测文件查询：
        1. 查询知识库下所有符合条件的评测文件（支持模糊匹配、分页）
        2. 每个文件返回：最后一条有效测评（状态3/4/5）的状态+消息、状态3/4/5的测评总条数
        """

        try:
            # 1. 构建文件查询条件（查询该知识库下所有符合条件的文件）
            query = db.query(
                KNOWLEDGE_EVALUATION.file_id,
                KNOWLEDGE_EVALUATION.file_name,
                KNOWLEDGE_EVALUATION.file_url,
                KNOWLEDGE_EVALUATION.create_time,
                KNOWLEDGE_EVALUATION.size
            ).filter(KNOWLEDGE_EVALUATION.kb_id == knowledge_id)

            # 文件名模糊匹配（可选）
            if file_name:
                query = query.filter(
                    KNOWLEDGE_EVALUATION.file_name.ilike(func.concat('%', file_name, '%'))
                )

            # 统计符合条件的文件总数（原始文件总数，未过滤测评状态）
            total = query.count()
            if total == 0:
                return {
                    "total": 0,
                    "total_pages": 0,
                    "page": page,
                    "page_size": page_size,
                    "result": []
                }

            # 分页处理（返回该页所有文件）
            offset = (page - 1) * page_size
            file_records = query.order_by(
                desc(KNOWLEDGE_EVALUATION.create_time)  # 文件按创建时间倒序排列
            ).offset(offset).limit(page_size).all()

            result_list = []
            for file in file_records:
                # 基础文件信息初始化
                file_info = {
                    "file_id": file.file_id,
                    "file_name": file.file_name or "",
                    "file_url": file.file_url or "",
                    "size": file.size or 0,
                    "file_create_time": file.create_time.strftime("%Y-%m-%d %H:%M:%S") if file.create_time else None,
                    # 核心字段1：状态3/4/5的测评总条数
                    "result_count": 0,
                    # 核心字段3：最后一条有效测评的消息（无则为空字符串）
                    "status":0,
                    "message": ""
                }

                # 2. 统计该文件下状态为3/4/5的测评总条数
                eval_count = db.query(KnowledgeEvaluationSetting).filter(
                    KnowledgeEvaluationSetting.file_id == file.file_id,
                    KnowledgeEvaluationSetting.status.in_([3, 4, 5])
                ).count()
                file_info["result_count"] = eval_count  # 赋值统计结果

                # 3. 查询当前文件最新的1条有效Setting记录（状态3/4/5，按时间倒序）
                latest_setting = db.query(
                    KnowledgeEvaluationSetting.status,
                    KnowledgeEvaluationSetting.message
                ).filter(
                    KnowledgeEvaluationSetting.file_id == file.file_id,
                ).order_by(
                    desc(KnowledgeEvaluationSetting.create_time)
                ).first()
                if latest_setting:
                    file_info["message"] = latest_setting.message or ""
                    file_info["status"] = latest_setting.status
                # 无论是否有测评记录，都保留该文件
                result_list.append(file_info)

            # 计算总页数（基于原始文件总数）
            total_pages = (total + page_size - 1) // page_size

            return {
                "total": total,  # 知识库下符合条件的文件总数
                "total_pages": total_pages,
                "page": page,
                "page_size": page_size,
                "result": result_list  # 仅包含需求字段
            }
        except Exception as e:
            logger.error(f"查询知识库[{knowledge_id}]评测文件失败：{str(e)}")
            raise

    @staticmethod
    def delete_single_evaluation(db: Session, evaluation_id: str) -> dict:
        """
        单个测评ID的精准删除逻辑
        步骤：
        1. 校验测评ID是否存在
        2. 按顺序删除：问题表→答案表→结果表→设置表
        3. 统计各表删除条数，确保事务原子性
        """
        try:
            # 步骤1：校验测评ID是否存在
            setting = db.query(KnowledgeEvaluationSetting).filter(
                KnowledgeEvaluationSetting.evaluation_id == evaluation_id
            ).first()
            if not setting:
                return {
                    "success": False,
                    "message": f"测评ID【{evaluation_id}】不存在，无需删除",
                    "delete_count": 0,
                }
            delete_count = 0
            ea_delete = db.query(EvaluationAnswer).filter(
                EvaluationAnswer.evaluation_id == evaluation_id
            ).delete(synchronize_session=False)
            delete_count += ea_delete
            # ========== 步骤4：删除「结果表」evaluation_result ==========
            er_delete = db.query(EvaluationResult).filter(
                EvaluationResult.evaluation_id == evaluation_id
            ).delete(synchronize_session=False)
            delete_count += er_delete

            # ========== 步骤5：删除「测评设置表」evaluation_setting ==========
            es_delete = db.query(KnowledgeEvaluationSetting).filter(
                KnowledgeEvaluationSetting.evaluation_id == evaluation_id
            ).delete(synchronize_session=False)
            delete_count += es_delete
            # 提交事务（原子性保证）
            db.commit()
            return {
                "success": True,
                "message": f"测评ID【{evaluation_id}】删除完成",
                "delete_count": delete_count,
            }

        except Exception as e:
            # 异常回滚，确保数据一致性
            db.rollback()
            logger.exception(f"单个测评ID删除失败（evaluation_id={evaluation_id}）：{str(e)}")
            raise Exception(f"数据库操作异常：{str(e)}")

    @staticmethod
    def abort_evaluation_task(db: Session, file_id: str) -> None:
        """
        中止评测任务（服务层）- 100% 保留原业务逻辑
        入参：仅接收db和file_id（无路由相关对象）
        异常：直接抛出，由路由层按原格式返回错误
        """
        task = db.query(KNOWLEDGE_EVALUATION).filter(KNOWLEDGE_EVALUATION.file_id == file_id).first()
        if not task:
            raise Exception(f"文件ID {file_id} 不存在")
        setting = db.query(KnowledgeEvaluationSetting).filter(
            KnowledgeEvaluationSetting.file_id == file_id,
            KnowledgeEvaluationSetting.status.in_([1, 2,3,4,5])
        ).order_by(KnowledgeEvaluationSetting.update_time.desc()).first()
        if not setting:
            raise Exception("未找到可中止的评测任务")
        task_id = f"{setting.evaluation_id}:evaluation"
        result = AsyncResult(task_id, app=celery_app)
        if result.state == "PENDING":
            result.revoke()
            logger.info(f"任务 {task_id} 取消成功（排队中）")
        elif result.state in ("RECEIVED", "STARTED"):
            result.revoke(terminate=True, signal="SIGTERM")
            logger.info(f"任务 {task_id} 取消成功（正在执行）")
        elif result.state in ("SUCCESS", "FAILURE", "REVOKED"):
            logger.info(f"任务 {task_id} 已完成，当前状态: {result.state}")
        else:
            result.revoke(terminate=True)
            logger.info(f"任务 {task_id} 处于未知状态: {result.state}，已尝试取消")
        db.query(KnowledgeEvaluationSetting).filter(
            KnowledgeEvaluationSetting.evaluation_id == setting.evaluation_id
        ).update({
            KnowledgeEvaluationSetting.status: 5,
            KnowledgeEvaluationSetting.message: "用户终止",
            KnowledgeEvaluationSetting.update_time: datetime.now().strftime("%Y%m%d%H%M%S")
        })
        db.commit()
        logger.info(f"评测任务已中止: file_id={file_id}")
        evaluation_id = setting.evaluation_id
        evaluation_results = db.query(EvaluationResult).filter(
            EvaluationResult.evaluation_id == evaluation_id
        ).all()

        if evaluation_results:
            for res in evaluation_results:
                db.delete(res)
            db.commit()
            logger.info(f"已删除evaluation_id={evaluation_id}对应的EvaluationResult数据，共{len(evaluation_results)}条")

        return
