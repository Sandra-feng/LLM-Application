import io
import mimetypes
import os
import tempfile
import time
import uuid  # 添加uuid模块导入
from pathlib import Path
from typing import Any
from xml.etree.ElementTree import fromstring
from zipfile import ZipFile

import chardet
import httpx
import openpyxl
import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.packaging.relationship import get_dependents, get_rels_path
from openpyxl.utils import range_boundaries
from openpyxl.xml.constants import IMAGE_NS, REL_NS, SHEET_DRAWING_NS
from PIL import Image as PILImage

from base_configs.api_config import ApiConfig
from base_configs.minio_config import MinioConfig
from base_configs.mongodb_config import CollectionConfig
from base_utils.minio_util import MinIoUtil
from base_utils.mongodb_util import MongodbUtil
from service_knowledge_manage.service.parsers.base_parser import BaseParser


class ExcelParser(BaseParser):
    """Excel 文件（.xls, .xlsx）和CSV文件解析器。"""

    def __init__(self):
        """初始化Excel解析器"""
        super().__init__()
        self.temp_dir = Path(__file__).resolve().parents[1] / "result" / "images"
        os.makedirs(self.temp_dir, exist_ok=True)

    def _convert_xls_to_xlsx(self, file_path: str) -> str:
        """将xls文件转换为xlsx格式"""
        try:
            import xlrd
            from openpyxl import Workbook

            # 读取xls文件
            xls_book = xlrd.open_workbook(file_path)

            # 创建临时xlsx文件
            temp_dir = tempfile.gettempdir()
            xlsx_path = os.path.join(temp_dir, f"{os.path.basename(file_path)}.xlsx")

            # 创建新的工作簿
            wb = Workbook()
            wb.remove(wb.active)  # 移除默认的工作表

            # 复制每个工作表
            for sheet_index in range(xls_book.nsheets):
                sheet = xlrd.open_workbook(file_path).sheet_by_index(sheet_index)
                ws = wb.create_sheet(title=sheet.name)

                # 复制数据
                for row in range(sheet.nrows):
                    for col in range(sheet.ncols):
                        try:
                            cell_value = sheet.cell_value(row, col)
                            ws.cell(row=row + 1, column=col + 1, value=cell_value)
                        except Exception as e:
                            logger.warning(f"复制单元格数据失败: {e}")
                            continue

            # 保存xlsx文件
            wb.save(xlsx_path)
            return xlsx_path
        except Exception as e:
            logger.info(f"xls转换xlsx失败: {e}")
            raise Exception(f"xls转换xlsx失败: {e}")

    async def unmerge_cells(self, ws_value, ws_formula, is_save_image=False):
        """拆分合并单元格，并根据 is_save_image 决定是否恢复合并"""
        try:
            logger.info("开始处理合并单元格")
            merge_ranges = list(ws_formula.merged_cells.ranges)
            logger.info(f"合并单元格区域列表：{merge_ranges}")
            # 记录所有合并区域，用于后续恢复
            original_merged_ranges = [str(merged_range) for merged_range in merge_ranges]
            for merged_range in list(ws_formula.merged_cells.ranges):
                min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                top_left_value = ws_value.cell(row=min_row, column=min_col).value
                ws_formula.unmerge_cells(str(merged_range))
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        ws_formula.cell(row=r, column=c).value = top_left_value
            # if not is_save_image:
            #     logger.info("is_save_image=False，恢复合并单元格")
            #     for merged_range in original_merged_ranges:
            #         ws_formula.merge_cells(merged_range)
            logger.info("合并单元格处理完成")
            return ws_formula
        except Exception as e:
            logger.info(f"合并单元格处理失败{str(e)}")
            raise Exception(f"合并单元格处理失败")

    async def unmerge_cells_v1(self, ws_value, ws_formula, is_save_image=False):
        """拆分合并单元格，并根据 is_save_image 决定是否恢复合"""
        try:
            logger.info("开始处理合并单元格")
            merge_ranges = list(ws_formula.merged_cells.ranges)
            logger.info(f"合并单元格区域列表：{merge_ranges}")
            # 记录所有合并区域，用于后续恢复
            original_merged_ranges = [str(merged_range) for merged_range in merge_ranges]
            for merged_range in list(ws_formula.merged_cells.ranges):
                min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                top_left_value = ws_value.cell(row=min_row, column=min_col).value
                ws_formula.unmerge_cells(str(merged_range))
                for r in range(min_row, max_row + 1):
                    for c in range(min_col, max_col + 1):
                        ws_formula.cell(row=r, column=c).value = top_left_value
            if not is_save_image:
                logger.info("is_save_image=False，恢复合并单元格")
                for merged_range in original_merged_ranges:
                    ws_formula.merge_cells(merged_range)
            logger.info("合并单元格处理完成")
            return ws_formula
        except Exception as e:
            logger.info("合并单元格处理失败", str(e))
            raise

    async def extract_images_from_excel(self, file_path):
        """提取图片"""
        try:
            # 检查 cellimages.xml 是否存在
            CELLIMAGE_PATH = "xl/cellimages.xml"
            archive = ZipFile(file_path, "r")

            if CELLIMAGE_PATH not in archive.namelist():
                logger.info("Excel文件中没有图片，跳过图片提取")
                archive.close()
                return {}

            def parse_element(element):
                data = {}
                xdr_namespace = "{%s}" % SHEET_DRAWING_NS
                targets = self.level_order_traversal(element, xdr_namespace + "nvPicPr")
                for target in targets:
                    cNvPr = embed = ""
                    for child in target:
                        if child.tag == xdr_namespace + "nvPicPr":
                            cNvPr = child[0].attrib["name"]
                        elif child.tag == xdr_namespace + "blipFill":
                            _rel_embed = "{%s}embed" % REL_NS
                            embed = child[0].attrib[_rel_embed]
                    if cNvPr:
                        data[cNvPr] = embed
                return data

            def handle_images(deps, archive):
                images = []
                for dep in deps:
                    if dep.Target == "NULL":
                        logger.info("图片格式不支持，跳过")
                        continue
                    if dep.Type != IMAGE_NS:
                        logger.info("图片格式不支持，跳过")
                        continue
                    try:
                        image_io = archive.read(dep.target)
                        image = XLImage(io.BytesIO(image_io))
                    except OSError:
                        logger.info("图片不可读，跳过")
                        continue
                    if hasattr(image, "format") and image.format.upper() == "WMF":
                        logger.info("图片无法保存，跳过")
                        continue
                    image.embed = dep.id
                    image.target = dep.target
                    images.append(image)
                return images

            src = archive.read(CELLIMAGE_PATH)
            deps = get_dependents(archive, get_rels_path(CELLIMAGE_PATH))
            image_rels = handle_images(deps=deps, archive=archive)
            node = fromstring(src)
            cellimages_xml = parse_element(node)
            cellimages_rel = {}
            for image in image_rels:
                cellimages_rel[image.embed] = image
            for cnvpr, embed in cellimages_xml.items():
                cellimages_xml[cnvpr] = cellimages_rel.get(embed)
            archive.close()
            return cellimages_xml
        except Exception as e:
            logger.info("提取Excel图片失败", str(e))
            raise Exception(f"提取Excel图片失败")

    async def overwrite_formula_cell(self, ws_value, ws_formula, is_save_image=False):
        """替换公式单元格为计算值，如果is_save_image为False则清除图片公式"""
        try:
            count = 0
            logger.info("开始处理单元格公式赋值")
            for row in ws_formula.iter_rows():
                for cell in row:
                    # 检查是否为公式单元格且值不为None
                    if cell.data_type == "f" and cell.value is not None:
                        cell_value_str = str(cell.value)
                        # 如果是图片相关的公式
                        if "DISPIMG" in cell_value_str.upper():
                            # 如果不需要保存图片，则清空单元格
                            if not is_save_image:
                                cell.value = ""
                                count += 1
                            # 如果需要保存图片，则保持原公式不变
                            continue
                        # 对于其他公式，获取计算值
                        calculated_value = ws_value[cell.coordinate].value
                        # 只有当计算值不为None时才进行替换
                        if calculated_value is not None:
                            cell.value = calculated_value
                        else:
                            # 如果计算值为None，保持原公式
                            logger.debug(f"单元格 {cell.coordinate} 的计算值为None，保持原公式")
            logger.info("单元格公式赋值处理完成")
            return ws_formula
        except Exception as e:
            logger.info("单元格公式赋值处理失败", str(e))
            raise Exception(f"单元格公式赋值处理失败")

    async def overwrite_newline_cell(self, df):
        """替换换行符"""
        try:
            logger.info("开始处理单元格换行符替换")
            df = df.map(lambda x: x.replace("\n", "").replace("\r", "") if isinstance(x, str) else x)
            logger.info("单元格换行符替换完成")
            return df
        except Exception as e:
            logger.info("单元格换行符替换失败", str(e))
            raise Exception(f"单元格换行符替换失败")

    async def to_dataframe(self, ws_formula):
        """转换为dataframe格式"""
        try:
            data = []
            for row in ws_formula.iter_rows(values_only=True):
                data.append(row)
            return pd.DataFrame(data)
        except Exception as e:
            logger.error("转换为DataFrame失败")
            raise Exception(f"转换为DataFrame失败:{e}")

    def level_order_traversal(self, root, flag):
        """层级遍历查找元素"""
        try:
            queue = [root]
            targets = []
            while queue:
                node = queue.pop(0)
                children = [child.tag for child in node]
                if flag in children:
                    targets.append(node)
                    continue
                for child in node:
                    queue.append(child)
            return targets
        except Exception as e:
            logger.error("层级遍历失败")
            raise Exception(f"层级遍历失败:{e}")

    async def get_image_description(self, image_path: str, **kwargs) -> str:
        """同步版本的图片描述获取"""
        try:
            if not os.path.exists(image_path):
                logger.warning(f"图片文件不存在: {image_path}")
                return "图片描述"

            # 获取参数
            use_vlm_mode = False
            model_uid = kwargs.get("model_uid", "Qwen2.5-VL-7B-Instruct")
            api_url = kwargs.get("api_url", "http://10.8.21.164:9997/v1")
            api_key = kwargs.get("api_key", "not empty")
            knowledge_id = kwargs.get("knowledge_id", "")

            if model_uid and api_url and api_key:
                use_vlm_mode = True

            # 构建表单数据
            data = {
                "use_vlm_mode": use_vlm_mode,
                "return_middle_json": "false",
                "return_model_output": "false",
                "return_md": "false",
                "return_images": "false",
                "end_page_id": "99999",
                "api_url": api_url,
                "parse_method": "auto",
                "start_page_id": "0",
                "lang_list": "ch",
                "output_dir": "./output",
                "use_public_url": "false",
                "minio_bucket": "tiance-base",
                "server_url": "string",
                "return_content_list": "true",
                "backend": "pipeline",
                "table_enable": "true",
                "model_uid": model_uid,
                "api_key": api_key,
                "knowledge_id": knowledge_id,
                "enable_minio": "false",
                "formula_enable": "true",
            }
            mime_type, _ = mimetypes.guess_type(image_path)
            if not mime_type:
                mime_type = "application/octet-stream"

            # 打开文件并发送请求
            with open(image_path, "rb") as f:
                files = {"files": (os.path.basename(image_path), f, mime_type)}

                async with httpx.AsyncClient(base_url=f"{ApiConfig.Small_Model_Address}", timeout=60.0) as client:
                    resp = await client.post("/file_parse", data=data, files=files)
                    resp.raise_for_status()

                    # 解析响应
                    result = resp.json()
                    # 从API响应中提取图片描述
                    if (
                        result.get("results")
                        and result["results"].get("content_list")
                        and len(result["results"]["content_list"]) > 0
                    ):
                        return result["results"]["content_list"][0].get("text", "图片描述")

                    return "图片描述"
        except Exception as e:
            logger.error(f"获取图片描述失败: {e}")
            return "获取图片描述失败"

    async def process_excel_images(self, df, image_rels, archive, sheet_name, **kwargs):
        """处理表格中的嵌入图片"""
        try:
            knowledge_id = kwargs.get("knowledge_id", "")
            rows_with_images = []
            for row_idx, row in df.iterrows():
                row_text_parts = []
                row_images = []
                for col_idx, value in enumerate(row):
                    if isinstance(value, str):
                        matched = False
                        for image_id, img_obj in image_rels.items():
                            if img_obj and image_id in value:
                                try:
                                    img_path = img_obj.target
                                    img_bytes = archive.read(img_path)
                                    suffix = Path(img_path).suffix
                                    # 使用PIL打开图片
                                    pil_img = PILImage.open(io.BytesIO(img_bytes))
                                    # 生成唯一文件名
                                    temp_id = f"img_{sheet_name}_{row_idx}_{col_idx}_{int(time.time())}"
                                    file_name = f"{temp_id}{suffix}"
                                    output_path = os.path.join(self.temp_dir, file_name)
                                    remote_path = f"{knowledge_id}/image/{file_name}"
                                    pil_img.save(output_path)
                                    MinIoUtil.upload_image_file(MinioConfig.BUCKET_NAME, remote_path, output_path)
                                    image_desc = await self.get_image_description(output_path, **kwargs)
                                    # 获取图片尺寸
                                    width, height = pil_img.size

                                    # 记录图片详细信息
                                    row_images.append(
                                        {
                                            "image_id": image_id,
                                            "id": temp_id,
                                            "path": output_path,
                                            "remote_path": remote_path,
                                            "desc": image_desc,
                                            "type": "embedded",
                                            "position": {
                                                "column": col_idx + 1,  # 列索引加1
                                                "row": row_idx + 1,  # 行索引加1
                                                "columnOffset": 0,
                                                "rowOffset": 0,
                                            },
                                            "width": width,
                                            "height": height,
                                        }
                                    )
                                    matched = True
                                except Exception as e:
                                    logger.error(f"处理图片失败: {e}")
                                    continue
                        if not matched:
                            row_text_parts.append(value)
                    else:
                        row_text_parts.append(str(value) if value is not None else "")
                rows_with_images.append(
                    {"row_index": row_idx, "text": " ".join(row_text_parts).strip(), "images": row_images}
                )
            return rows_with_images
        except Exception as e:
            logger.error("表格嵌入图片处理失败")
            raise Exception(f"表格嵌入图片处理失败:{e}")

    async def process_floating_images(self, df, ws, archive, sheet_name, **kwargs):
        """处理表格中的浮动图片"""
        try:
            rows_with_images = []
            n_rows = len(df)  # 获取DataFrame的行数
            knowledge_id = kwargs.get("knowledge_id", "")
            # 确保图片目录存在
            os.makedirs(self.temp_dir, exist_ok=True)

            for img_idx, img in enumerate(ws._images):
                try:
                    logger.info(f"处理第 {img_idx + 1} 个浮动图片")

                    # 获取图片锚点信息
                    anchor = img.anchor._from

                    # 直接使用PIL处理图片
                    try:
                        # 从图片对象获取数据
                        if hasattr(img, "ref") and img.ref:
                            from PIL import Image

                            # 使用PIL打开图片
                            pil_img = Image.open(img.ref).convert("RGB")

                            # 获取图片尺寸
                            width, height = pil_img.size
                            logger.info(f"图片尺寸: {width}x{height}")

                            # 生成唯一文件名
                            temp_id = f"float_img_{sheet_name}_{img_idx}_{int(time.time())}"
                            file_name = f"{temp_id}.jpg"  # 保存为JPG格式
                            output_path = os.path.join(self.temp_dir, file_name)
                            remote_path = f"{knowledge_id}/image/{file_name}"

                            # 保存图片到临时文件
                            pil_img.save(output_path)
                            logger.info(f"图片已保存到临时文件: {output_path}")

                            # 获取图片描述
                            try:
                                image_desc = await self.get_image_description(output_path, **kwargs)
                                logger.info(f"图片描述: {image_desc}")
                            except Exception as e:
                                logger.error(f"获取图片描述失败: {e}")
                                image_desc = "图片描述"

                            # 上传到MinIO
                            try:
                                MinIoUtil.upload_image_file(MinioConfig.BUCKET_NAME, remote_path, output_path)
                                logger.info(f"图片已上传到MinIO: {remote_path}")
                            except Exception as e:
                                logger.error(f"上传图片到MinIO失败: {e}")
                                continue
                            col_offset = getattr(anchor, "colOff", 0) if hasattr(anchor, "colOff") else 0
                            row_offset = getattr(anchor, "rowOff", 0) if hasattr(anchor, "rowOff") else 0
                            from_row, from_col = anchor.row + 1, anchor.col + 1
                            to_row, to_col = img.anchor.to.row + 1, img.anchor.to.col + 1
                            logger.info(f"浮动图片覆盖区域：起始{from_row}行{from_col}列，终止{to_row}行{to_col}列")
                            row_idx = from_row - 1  # 转换为0-based索引
                            if row_idx >= n_rows:
                                logger.warning(f"行索引 {row_idx} 超出DataFrame范围 (0-{n_rows - 1})，跳过该图片")
                                continue
                            try:
                                row_text = " ".join([str(v) for v in df.iloc[row_idx].tolist() if v is not None])
                                rows_with_images.append(
                                    {
                                        "row_index": row_idx,
                                        "text": row_text,
                                        "images": [
                                            {
                                                "id": temp_id,
                                                "path": output_path,
                                                "remote_path": remote_path,
                                                "desc": image_desc,
                                                "type": "floating",
                                                "position": {
                                                    "column": from_col,
                                                    "row": from_row,
                                                    "columnOffset": col_offset,
                                                    "rowOffset": row_offset,
                                                },
                                                "width": width,
                                                "height": height,
                                                "range": {
                                                    "from_row": from_row,
                                                    "from_col": from_col,
                                                    "to_row": to_row,
                                                    "to_col": to_col,
                                                },
                                                "rotation": 0,  # 默认值
                                                "scale_x": 1.0,  # 默认值
                                                "scale_y": 1.0,  # 默认值
                                            }
                                        ],
                                    }
                                )
                                logger.info(f"为第 {row_idx + 1} 行添加浮动图片记录")
                            except Exception as e:
                                logger.error(f"处理浮动图片行失败: {e}")
                                continue
                        else:
                            logger.warning("图片没有有效的引用，跳过处理")
                    except Exception as e:
                        logger.error(f"使用PIL处理图片失败: {e}")
                        continue

                except Exception as e:
                    logger.error(f"处理浮动图片失败: {e}")
                    continue

            logger.info(f"浮动图片处理完成，共处理 {len(rows_with_images)} 行")
            return rows_with_images
        except Exception as e:
            logger.error("浮动图片处理失败")
            raise Exception(f"浮动图片处理失败:{e}")

    async def header_merge(self, df, is_header_config, start_lines, end_lines, merge_method):
        """表头拼接"""
        try:
            if is_header_config:
                if end_lines > 1:
                    new_headers = []
                    if merge_method == "all":
                        headers = df.iloc[start_lines - 1 : end_lines, :]
                        columns = headers.shape[1]
                        header_df_str = headers.astype(str)
                        for column in range(columns):
                            header_merge = "_".join(list(header_df_str.values[:, column]))
                            new_headers.append(header_merge)
                    else:
                        headers = df.iloc[end_lines - 1, :]
                        header_df_str = headers.astype(str)
                        new_headers = list(header_df_str.values)
                else:
                    headers = df.iloc[0, :]
                    header_df_str = headers.astype(str)
                    new_headers = list(header_df_str.values)
                df.columns = new_headers
                df = df.iloc[end_lines:, :]
            return df
        except Exception as e:
            logger.error("表头拼接失败")
            raise Exception(f"表头拼接失败:{e}")

    async def content_merge(self, df, is_content_merge, image_rels):
        """内容拼接"""
        try:
            result = []
            df = df.astype(str)
            columns = list(df.columns)

            def replace_image_with_describe(value):
                if isinstance(value, str):
                    for image_id, img_obj in image_rels.items():
                        if img_obj and image_id in value:
                            value = ""
                return value

            df = df.apply(lambda col: col.map(replace_image_with_describe))
            if is_content_merge:
                for data in df.values:
                    result.append("，".join([f"{columns[i]}：{data[i]}" for i in range(len(data))]))
            else:
                result = ["，".join(i) for i in df.values]
            return result
        except Exception as e:
            logger.error("内容拼接失败")
            raise Exception(f"内容拼接失败:{e}")

    async def merge_rows_with_images(self, embed_rows, float_rows):
        """图片列融合"""
        try:
            merged = {}
            for row in embed_rows:
                if row["images"]:
                    merged[row["row_index"]] = row
            for row in float_rows:
                idx = row["row_index"]
                if idx in merged:
                    merged[idx]["images"].extend(row["images"])
                else:
                    if row["images"]:
                        merged[idx] = row
            merged_list = sorted(merged.values(), key=lambda x: x["row_index"])
            return merged_list
        except Exception as e:
            logger.error("图片列融合失败")
            raise Exception(f"图片列融合失败:{e}")

    # 新增方法：生成随机UUID作为节点ID
    async def generate_node_id(self) -> str:
        return str(uuid.uuid4())

    # 新增方法：将最终结果转换为节点
    async def convert_final_result_to_nodes(self, final_result: list[dict]) -> list[dict]:
        """将最终结果转换为标准节点格式"""
        try:
            nodes = []
            for item in final_result:
                chunk_text = item.get("chunk_content", "")
                images = item.get("source_data", {}).get("content", [])
                urls = item.get("source_data", {}).get("source_url", [])
                line = item.get("result_list", [])
                # 创建文本节点
                text_node = {
                    "type": "text",
                    "text": chunk_text,
                    "page_idx": 0,
                    "id": await self.generate_node_id(),  # 使用随机UUID
                    "referenced_images": [],
                    "referenced_tables": [],
                }
                nodes.append(text_node)
                # 创建图片节点（如果有图片）
                if images and urls:
                    for desc, url in zip(images, urls):
                        image_node = {
                            "type": "image",
                            "text": desc,  # 图片描述作为文本
                            "page_idx": 0,
                            "id": await self.generate_node_id(),  # 使用随机UUID
                            "referenced_images": [url],  # 图片URL
                            "referenced_tables": [],
                        }
                        nodes.append(image_node)
            return nodes
        except Exception as e:
            logger.error("转换最终结果为节点失败")
            raise Exception(f"转换最终结果为节点失败:{e}")

    async def _detect_file_encoding(self, file_path: str, sample_size=100000):
        """检测文件编码"""
        try:
            logger.info("开始检测文件编码")
            with open(file_path, "rb") as f:
                raw_data = f.read(sample_size)
                result = chardet.detect(raw_data)
                encoding = result["encoding"]
                confidence = result["confidence"]
                return encoding, confidence
        except Exception as e:
            logger.error(f"文件编码检测失败: {e}")
            raise Exception(f"文件编码检测失败:{e}")

    async def _read_csv_with_encoding(self, file_path: str, encoding: str = None):
        """读取CSV文件"""
        try:
            logger.info("开始读取CSV文件")
            # 自动检测编码
            detected_encoding, confidence = await self._detect_file_encoding(file_path)
            if detected_encoding and confidence > 0.7:
                try:
                    df = pd.read_csv(
                        file_path,
                        header=None,
                        dtype=str,
                        keep_default_na=False,
                        na_values=[""],
                        encoding=detected_encoding,
                    )
                    logger.info(f"使用检测到的编码 '{detected_encoding}' 成功读取文件")
                    df = df.replace({'\n': ' ', '\r': ' '}, regex=True)
                    return df
                except (UnicodeDecodeError, LookupError) as e:
                    logger.warning(f"检测到的编码 '{detected_encoding}' 读取失败: {e}")
            # 尝试使用utf-8编码读取
            try:
                df = pd.read_csv(
                    file_path, header=None, dtype=str, keep_default_na=False, na_values=[""], encoding="utf-8"
                )
                logger.info("使用utf-8编码成功读取文件")
                return df
            except (UnicodeDecodeError, LookupError) as e:
                logger.warning(f"utf-8编码读取失败: {e}")
            # 尝试使用gbk编码读取
            try:
                df = pd.read_csv(
                    file_path, header=None, dtype=str, keep_default_na=False, na_values=[""], encoding="gbk"
                )
                logger.info("使用gbk编码成功读取文件")
                return df
            except (UnicodeDecodeError, LookupError) as e:
                logger.warning(f"gbk编码读取失败: {e}")
            # 如果所有编码都失败，抛出异常
            raise Exception("无法确定文件编码，所有尝试的编码格式均失败")
        except Exception as e:
            logger.error("CSV文件读取失败")
            raise Exception(f"CSV文件读取失败:{e}")

    async def _build_final_columns(self, header_rows_matrix, merge_method):
        """构建最终列名"""
        try:
            logger.info("开始构建最终列名")
            if not header_rows_matrix:
                return []

            ncols = max(len(r) for r in header_rows_matrix)
            final_cols = []

            for c in range(ncols):
                parts = []
                for r in range(len(header_rows_matrix)):
                    # 确保当前行有足够的列
                    if c < len(header_rows_matrix[r]):
                        cell_value = str(header_rows_matrix[r][c]).strip()
                        # 即使为空也保留，但跳过完全为空的情况
                        if cell_value:
                            parts.append(cell_value)
                        else:
                            # 对于空单元格，添加一个空占位符
                            parts.append("")
                    else:
                        parts.append("")

                # 根据合并方法处理列名
                if merge_method == "all":
                    # 连接所有非空部分
                    non_empty_parts = [p for p in parts if p]
                    colname = "_".join(non_empty_parts) if non_empty_parts else " "
                else:  # "last_only"
                    # 从后往前找第一个非空值
                    for i in range(len(parts) - 1, -1, -1):
                        if parts[i]:
                            colname = parts[i]
                            break
                    else:
                        colname = ""

                final_cols.append(colname)

            # 处理重复列名
            # seen = {}
            # for i, name in enumerate(final_cols):
            #     if name in seen:
            #         seen[name] += 1
            #         final_cols[i] = f"{name}"
            #     else:
            #         seen[name] = 0

            logger.info(f"构建的最终列名: {final_cols}")
            return final_cols
        except Exception as e:
            logger.error("构建最终列名失败")
            raise Exception(f"构建最终列名失败:{e}")

    async def _semantic_colnames_from_data(self, df: pd.DataFrame):
        """从数据中推断语义列名 - 简化版本，全部返回空字符串"""
        try:
            logger.info("开始创建空列名")
            # 直接返回与列数相同数量的空字符串
            colnames = [""] * df.shape[1]
            logger.info(f"创建的列名: {colnames}")
            return colnames
        except Exception as e:
            logger.error("创建列名失败")
            raise Exception(f"创建列名失败:{e}")

    async def _csv_header_merge(self, df, is_header_config, start_lines, end_lines, merge_method):
        """CSV表头拼接"""
        try:
            logger.info("开始处理表头拼接")
            if is_header_config:
                if end_lines > 1:
                    if merge_method == "all":
                        headers = df.iloc[start_lines - 1 : end_lines, :]
                        columns = headers.shape[1]
                        header_df_str = headers.astype(str)
                        new_headers = []
                        for column in range(columns):
                            header_merge = "_".join(list(header_df_str.values[:, column]))
                            new_headers.append(header_merge)
                    else:  # "last_only"
                        headers = df.iloc[end_lines - 1, :]
                        header_df_str = headers.astype(str)
                        new_headers = list(header_df_str.values)
                else:
                    headers = df.iloc[0, :]
                    header_df_str = headers.astype(str)
                    new_headers = list(header_df_str.values)
                df.columns = new_headers
                df = df.iloc[end_lines:, :]
            logger.info("表头拼接处理完成")
            return df
        except Exception as e:
            logger.error("表头拼接处理失败")
            raise Exception(f"表头拼接处理失败:{e}")

    async def _csv_content_merge(self, df, is_content_merge):
        """CSV内容拼接"""
        try:
            logger.info("开始处理内容拼接")
            result = []
            df = df.astype(str)
            columns = list(df.columns)
            if is_content_merge:
                for data in df.values:
                    result.append("，".join([f"{columns[i]}：{data[i]}" for i in range(len(data))]))
            else:
                result = ["，".join(i) for i in df.values]
            logger.info("内容拼接处理完成")
            return result
        except Exception as e:
            logger.error("内容拼接处理失败")
            raise Exception(f"内容拼接处理失败:{e}")

    async def _parse_csv(self, file_path: str, **kwargs) -> dict[str, Any]:
        """解析CSV文件"""
        file_id = kwargs.get("file_id", "")
        try:
            # 获取参数
            is_header_config = kwargs.get("is_header_config", False)
            start_line = kwargs.get("start_line", 1)  # 注意参数名改为单数形式
            end_line = kwargs.get("end_line", 1)  # 注意参数名改为单数形式
            merge_method = kwargs.get("header_merge_method", "all")
            is_content_merge = kwargs.get("is_content_merge", True)
            encoding = kwargs.get("encoding")
            file_id = kwargs.get("file_id", "")

            # 读取CSV文件
            df_all = await self._read_csv_with_encoding(file_path, encoding)
            df_all = df_all.fillna("")
            if df_all.empty:
                raise Exception("CSV文件读取失败，文件可能为空或格式不正确")
            total_rows = df_all.shape[0]

            # 处理表头
            if is_header_config:
                # 确保行号在有效范围内
                if start_line > total_rows:
                    start_line = 1
                if end_line > total_rows or end_line < start_line:
                    end_line = start_line
                # 提取表头行
                header_matrix = []
                for r in range(start_line - 1, end_line):
                    if r < len(df_all):
                        row_data = df_all.iloc[r].fillna("").astype(str).tolist()
                        header_matrix.append(row_data)
                final_cols = await self._build_final_columns(header_matrix, merge_method)

                if not final_cols:
                    final_cols = ""
                # 提取数据行
                data_df = df_all.iloc[end_line:].reset_index(drop=True)
            else:
                # 无表头模式，自动推断列名
                final_cols = await self._semantic_colnames_from_data(df_all)
                data_df = df_all.copy().reset_index(drop=True)

            # 确保数据 DataFrame 有足够的列
            if data_df.shape[1] < len(final_cols):
                # 添加缺失的列
                for i in range(len(final_cols) - data_df.shape[1]):
                    data_df[data_df.shape[1] + i] = ""
            elif data_df.shape[1] > len(final_cols):
                # 截断多余的列
                data_df = data_df.iloc[:, : len(final_cols)]

            # 设置列名
            data_df.columns = final_cols

            # 处理内容拼接，获取每行的文本内容
            content_rows = await self._csv_content_merge(data_df, is_content_merge)

            # 按行创建节点
            nodes = []
            for row_idx, row_text in enumerate(content_rows, start=1):
                node = {
                    "type": "text",
                    "text": row_text,
                    "actual_row_idx": row_idx,
                    "page_idx": 0,
                    "bbox": [],
                    "id": await self.generate_node_id(),
                    "referenced_images": [],
                    "referenced_tables": [],
                }
                nodes.append(node)

            return {
                "status": "Success",
                "knowledge_id": kwargs.get("knowledge_id", ""),
                "results": {
                    "content_list": {
                        "result": nodes,
                        "metadata": {"file_type": "csv", "file_id": file_id, "row_count": total_rows},
                    }
                },
            }

        except Exception as e:
            logger.error(f"CSV文件解析失败: {e}")
            raise Exception(f"CSV文件解析失败:{e}")

    async def _parse_excel(self, file_path: str, **kwargs) -> dict[str, Any]:
        """解析Excel文件"""
        file_id = kwargs.get("file_id", "")
        try:
            # 获取是否保存图片参数
            is_save_image = kwargs.get("is_save_image", False)
            # 提取图片（只有当需要保存图片时才执行）
            if is_save_image:
                try:
                    image_rels = await self.extract_images_from_excel(file_path)
                except Exception as e:
                    raise e
            else:
                image_rels = {}
                logger.info("is_save_image为False，跳过图片提取")
            archive = ZipFile(file_path, "r")
            wb_value = load_workbook(file_path, data_only=True)
            wb_formula = load_workbook(file_path, data_only=False)
            knowledge_id = kwargs.get("knowledge_id", "")
            # 初始化节点列表
            all_nodes = []
            if not is_save_image:
                try:
                    ws_value = wb_value[wb_value.sheetnames[0]]
                    ws_formula = wb_formula[wb_value.sheetnames[0]]
                    logger.info(f"开始处理工作表：{wb_value.sheetnames[0]}")
                    ws_formula = await self.unmerge_cells_v1(ws_value, ws_formula, is_save_image)
                    ws_formula = await self.overwrite_formula_cell(ws_value, ws_formula, is_save_image)
                    for row in ws_formula.iter_rows():
                        for cell in row:
                            try:
                                # 尝试设置值，如果失败说明是只读单元格
                                if cell.value is None:
                                    cell.value = ""
                            except AttributeError as e:
                                if "read-only" in str(e):
                                    continue  # 跳过只读单元格
                    wb_formula.save(file_path)
                    logger.info(f"已保存修改后的Excel文件: {file_path}")
                    # 获取文件名
                    if file_id:
                        file_info = MongodbUtil.query_doc_by_id(CollectionConfig.UPLOAD_FILE_INFO_COLLECTION, file_id)
                        file_name = file_info["file_name"]
                    else:
                        file_name = os.path.basename(file_path)
                    # 上传到MinIO
                    remote_path = f"{knowledge_id}/excel/{file_name}"
                    MinIoUtil.upload_file(MinioConfig.BUCKET_NAME, remote_path, file_path)
                    MongodbUtil.update_docs_by_condition(
                        "upload_file_info", {"_id": file_id}, replace_data={"$set": {"remove_image_path": remote_path}}
                    )
                except Exception as e:
                    logger.error(f"保存并上传修改后的Excel文件失败: {e}")
            wb_value = load_workbook(file_path, data_only=True)
            wb_formula = load_workbook(file_path, data_only=False)
            if not wb_value.sheetnames:
                archive.close()
                raise Exception(f"读取工作表 {wb_value.sheetnames}失败")
            sheet_name_index = 0
            try:
                ws_value = wb_value[wb_value.sheetnames[sheet_name_index]]
                ws_formula = wb_formula[wb_value.sheetnames[sheet_name_index]]
                logger.info(f"开始处理工作表：{wb_value.sheetnames[sheet_name_index]}")
                ws_formula = await self.unmerge_cells(ws_value, ws_formula, is_save_image)
                ws_formula = await self.overwrite_formula_cell(ws_value, ws_formula, is_save_image)
                df = await self.to_dataframe(ws_formula)
                data = pd.read_excel(file_path, header=None)
                if df.shape[0] > data.shape[0]:
                    df = df.iloc[: data.shape[0], :]
                # 2) 直接截断多余的列（如果有）
                if df.shape[1] > data.shape[1]:
                    df = df.iloc[:, : data.shape[1]]
                if df.empty:
                    raise Exception(f"处理工作表 {wb_value.sheetnames[sheet_name_index]} 失败: {e}")
                df = await self.overwrite_newline_cell(df)
                df = df.fillna("")
                # 获取参数
                is_header_config = kwargs.get("is_header_config", False)
                start_lines = kwargs.get("start_line", 1)
                end_lines = kwargs.get("end_line", 1)
                merge_method = kwargs.get("header_merge_method", "all")
                is_content_merge = kwargs.get("is_content_merge", True)
                df_with_header = await self.header_merge(
                    df,
                    is_header_config=is_header_config,
                    start_lines=start_lines,
                    end_lines=end_lines,
                    merge_method=merge_method,
                )
                df = await self.to_dataframe(ws_formula)
                df = df.fillna("")
                # 获取行高和列宽信息
                row_heights = {}
                column_widths = {}

                # 获取行高
                for row_idx in range(1, ws_formula.max_row + 1):
                    row_dim = ws_formula.row_dimensions[row_idx]
                    row_heights[str(row_idx)] = (
                        row_dim.height if row_dim.height else ws_formula.sheet_format.defaultRowHeight
                    )  # 默认行高

                # 获取列宽
                for col_idx in range(1, ws_formula.max_column + 1):
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    col_dim = ws_formula.column_dimensions[col_letter]
                    column_widths[str(col_idx)] = (
                        col_dim.width if col_dim.width else ws_formula.sheet_format.defaultColWidth
                    )  # 默认列宽
                content_rows = await self.content_merge(df_with_header, is_content_merge, image_rels)
                if is_save_image:
                    df_with_embed_image_describe = await self.process_excel_images(
                        df, image_rels, archive, wb_value.sheetnames[sheet_name_index], **kwargs
                    )
                    df_with_float_image_describe = await self.process_floating_images(
                        df, ws_formula, archive, wb_value.sheetnames[sheet_name_index], **kwargs
                    )
                    df_with_image_describe = await self.merge_rows_with_images(
                        df_with_embed_image_describe, df_with_float_image_describe
                    )
                else:
                    df_with_image_describe = []
                    logger.info("is_save_image为False，跳过图片处理")

                # 创建行索引到图片描述的映射
                row_images_map = {}
                for row_data in df_with_image_describe:
                    row_index = row_data["row_index"]
                    row_images_map[row_index] = row_data["images"]

                image_row_indices = set(row_images_map.keys())
                max_row_index = max(len(content_rows) - 1, max(image_row_indices) if image_row_indices else 0)
                while len(content_rows) <= max_row_index:
                    content_rows.append("")

                sheet_nodes = []
                content_rows = [row for row in content_rows if row.strip() != ""]
                embed_id_dict = {}
                for row_idx, row_text in enumerate(content_rows):
                    actual_row_idx = row_idx + end_lines if is_header_config else row_idx
                    # 获取该行的列宽信息
                    row_column_widths = {}
                    for col_idx in range(1, len(df_with_header.columns) + 1):
                        row_column_widths[col_idx] = column_widths.get(str(col_idx), 8.43)

                    image_nodes_in_this_row = []  # 用于保存该行的图片节点
                    image_ids_in_this_row = []  # 用于保存该行的图片节点ID
                    # 如果该行有图片且需要保存图片，则先创建图片节点
                    if is_save_image and actual_row_idx in row_images_map:
                        for img in row_images_map[actual_row_idx]:
                            image_node_id = await self.generate_node_id()
                            image_node = {
                                "type": "image",
                                "text": img["desc"],
                                "page_idx": 0,
                                "bbox": [],
                                "caption": [""],
                                "actual_row_idx": actual_row_idx,
                                "id": image_node_id,
                                "img_path": img["remote_path"],
                                "referenced_images": [],
                                "referenced_tables": [],
                                "image_info": {
                                    "image_id": img.get("image_id", ""),
                                    "image_type": img["type"],
                                    "position": img.get("position", {}),
                                    "width": img.get("width", 0),
                                    "height": img.get("height", 0),
                                    "range": img.get("range", {}),
                                },
                            }
                            if img["type"] == "embedded" and img["image_id"] in embed_id_dict:
                                if embed_id_dict[img["image_id"]] not in image_ids_in_this_row:
                                    image_ids_in_this_row.append(embed_id_dict[img["image_id"]])
                                continue
                            elif img["type"] == "embedded":
                                embed_id_dict[img["image_id"]] = image_node_id
                            image_ids_in_this_row.append(image_node_id)
                            image_nodes_in_this_row.append(image_node)

                    # 创建文本节点，并将该行的图片节点ID添加到referenced_images
                    text_node = {
                        "type": "text",
                        "text": row_text,
                        "actual_row_idx": actual_row_idx,
                        "bbox": [],
                        "page_idx": 0,
                        "id": await self.generate_node_id(),
                        "referenced_images": image_ids_in_this_row,  # 这里挂载图片节点的ID
                        "referenced_tables": [],
                    }
                    # 先添加文本节点
                    sheet_nodes.append(text_node)
                    # 再添加该行的图片节点
                    sheet_nodes.extend(image_nodes_in_this_row)
                all_nodes.extend(sheet_nodes)
            except Exception as e:
                logger.error(f"处理工作表 {wb_value.sheetnames[sheet_name_index]} 失败: {e}")
                # 继续处理下一个工作表
                raise Exception(f"处理工作表 {wb_value.sheetnames[sheet_name_index]} 失败: {e}")
            finally:
                archive.close()
            result = {
                "status": "Success",
                "knowledge_id": kwargs.get("knowledge_id", ""),
                "results": {
                    "content_list": {
                        "result": all_nodes,
                        "metadata": {
                            "file_type": "excel",
                            "file_id": file_id,
                            "row_heights": row_heights,  # 添加行高信息
                            "column_widths": column_widths,
                        },
                    }
                },
            }
            return result

        except Exception as e:
            logger.error("Excel文件解析失败")
            # 更新数据库状态
            raise Exception(f"Excel文件解析失败:{e}")

    async def parse(self, file_path: str, **kwargs) -> dict[str, Any]:
        """
        解析 Excel 或 CSV 文件。

        Args:
            file_path (str): 文件路径。
            **kwargs: 其他参数，包括解析选项。

        Returns:
            一个包含文本、元数据和切块的字典。
        """
        try:
            ext = os.path.splitext(file_path)[-1].lower()
            file_stem = Path(file_path).stem  # 获取文件名（不带扩展名）

            # 处理xls文件，先转换为xlsx
            if ext == ".xls":
                converted_path = self._convert_xls_to_xlsx(file_path)
                try:
                    kwargs["is_save_image"] = False
                    result = await self._parse_excel(converted_path, file_stem=file_stem, **kwargs)
                    # 删除临时文件
                    try:
                        os.remove(converted_path)
                    except:
                        pass
                    return result
                except Exception as e:
                    # 删除临时文件
                    try:
                        os.remove(converted_path)
                    except:
                        pass
                    raise Exception(f"xls解析失败:{e}")
            elif ext in [".xlsx"]:
                return await self._parse_excel(file_path, file_stem=file_stem, **kwargs)
            elif ext == ".csv":
                return await self._parse_csv(file_path, file_stem=file_stem, **kwargs)
            else:
                raise Exception("不支持当前文件格式")
        except Exception as e:
            logger.error(f"解析文件 {file_path} 失败: {e}")
            raise e
